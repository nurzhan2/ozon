# -*- coding: utf-8 -*-
"""
Импорт выгрузок из личного кабинета OZON.

ЗАЧЕМ
    Четыре строки отчёта «Качественные показатели» — показы, корзина,
    % корзины и место в поиске — через API без подписки Premium Plus взять
    негде. Проверены все четыре пути: /v1/analytics/data режет метрики по
    тарифу, /v1/analytics/product-queries отвечает через раз и корзины не
    знает вовсе, /v1/search-queries требует Premium Pro и отдаёт рыночную
    статистику по запросу, а не по нашему товару, в /v1/report/* нужного
    типа отчёта нет.

    В кабинете эти цифры есть и выгружаются файлом. Поэтому: тот, у кого
    есть доступ к кабинету, раз в день кладёт выгрузку в общую папку
    Google Диска, а сбор её читает и заполняет строки. Пароли нигде не
    хранятся, автоматизация браузера не нужна.

ЧТО ЧИТАЕМ
    xlsx и csv. Колонки ищутся по названиям, а не по номерам: OZON меняет
    порядок и формулировки, жёсткая привязка к позиции сломается на первой
    же правке. Нужны колонка с датой, колонка с артикулом или SKU и хотя бы
    одна из числовых.

    Если колонки с датой нет — файл пропускается с внятным сообщением.
    Выгрузка без разбивки по дням для подневного отчёта бесполезна, а
    размазать период по дням значило бы выдумать числа.

ГДЕ ИЩЕМ
    1. Папка на Google Диске (GOOGLE_IMPORT_FOLDER) с подпапками по имени
       магазина — так один каталог обслуживает все пять.
    2. Локально: DATA_DIR/import/<Магазин>/ — для прогонов на своей машине.

    Берётся самый свежий файл в папке магазина.
"""

import io
import os
import csv
import re
import logging
from datetime import datetime

log = logging.getLogger("ozon.cabinet")

# Названия колонок, которые встречаются в выгрузках кабинета. Список
# намеренно широкий: у разных отчётов OZON формулировки расходятся
# («Показы» / «Показы всего» / «Показы в поиске и категории»).
COLUMNS = {
    "day": ("дата", "день", "date", "period", "период"),
    "offer_id": ("артикул", "ваш sku", "offer id", "offer_id", "код товара"),
    "sku": ("sku", "ozon id", "озон id", "ид товара", "идентификатор товара"),
    "views": ("показы", "показов", "показы всего", "просмотры", "hits_view",
              "views"),
    # «Уникальные посетители, всего» — это session_view в терминах API:
    # уникальные люди, а не показы. Идёт в строку «клики».
    "sessions": ("уникальные посетители, всего", "сессии", "сессий",
                 "уникальные посетители", "посетители", "session_view",
                 "клики", "переходы"),
    # Посетители именно карточки: нужны, чтобы развернуть конверсию в штуки.
    "sessions_pdp": ("уникальные посетители с просмотром карточки товара",
                     "с просмотром карточки", "session_view_pdp"),
    "tocart": ("в корзину", "корзина", "добавления в корзину",
               "добавлено в корзину", "hits_tocart", "add_to_cart"),
    # В выгрузке кабинета корзина часто есть только процентом.
    "conv_tocart": ("конверсия в корзину из карточки товара",
                    "конверсия в корзину", "conv_tocart"),
    "position": ("позиция", "место в поиске", "средняя позиция",
                 "position", "position_category"),
}

# --- второй тип файла: выгрузка ЗАКАЗОВ ---
# Из неё берутся продажи по КЛАСТЕРУ ДОСТАВКИ. Это принципиально другая
# величина, чем ads_cluster из API: тот привязан к кластеру, где товар лежит,
# то есть к отгрузке. Заказчик считает потребность по доставке — куда товар
# уехал, а не откуда. В API кластера доставки нет ни в одном методе, только
# в этой выгрузке.
ORDER_COLUMNS = {
    "day": ("принят в обработку", "дата отгрузки", "дата", "день"),
    "offer_id": ("артикул", "ваш sku", "offer id"),
    "sku": ("sku", "ozon id"),
    "cluster": ("кластер доставки",),
    "cluster_ship": ("кластер отгрузки",),
    "qty": ("количество", "кол-во", "шт"),
    "status": ("статус",),
}
ORDER_ORDER = ("day", "cluster", "cluster_ship", "offer_id", "sku", "qty",
               "status")

# Отменённые заказы в продажи не идут. Заказчик их убирает руками ещё в
# кабинете («убрала отмены»), но полагаться на это нельзя.
CANCELLED = ("отмен", "cancel")

# Колонки, ради которых всё затевается. Файл без единой из них бесполезен.
VALUE_KEYS = ("views", "sessions", "sessions_pdp", "tocart", "conv_tocart",
              "position")

# Складываются при повторах; позиция и конверсия — нет, это доли и места.
SUM_KEYS = ("views", "sessions", "sessions_pdp", "tocart")


class CabinetImportError(Exception):
    pass


# ------------------------------------------------------------------ разбор

def _norm(s):
    """Название колонки к сравнимому виду: без регистра, пробелов и скобок."""
    s = str(s or "").strip().lower().replace("ё", "е")
    s = re.sub(r"\(.*?\)", " ", s)
    return re.sub(r"[\s ]+", " ", s).strip()


# Порядок разбора шапки. Роли идут от самых узких к широким, и одна колонка
# достаётся только одной роли. Без этого «Конверсия в корзину из карточки
# товара» попадала и в conv_tocart, и в tocart — потому что содержит слова
# «в корзину», — и в отчёт вместо штук уезжала доля 0,37.
ROLE_ORDER = ("day", "offer_id", "sku", "conv_tocart", "sessions_pdp",
              "tocart", "views", "sessions", "position")


def _match_columns(header):
    """{наша_роль: индекс_колонки}. Точное совпадение важнее вхождения."""
    norm = [_norm(h) for h in header]
    found, used = {}, set()
    for role in ROLE_ORDER:
        names = COLUMNS[role]
        for i, h in enumerate(norm):          # сначала точное совпадение
            if i not in used and h in names:
                found[role], _ = i, used.add(i)
                break
        if role in found:
            continue
        for i, h in enumerate(norm):          # затем вхождение
            if i not in used and any(n in h for n in names):
                found[role], _ = i, used.add(i)
                break
    return found


def _num(v):
    """
    «1 234,5», «1,234.5» и «36,99%» — в число. Мусор — в ноль.

    Процент возвращается долей: 36,99% -> 0.3699. Иначе конверсия, которой
    в выгрузке кабинета заменяют абсолютную корзину, была бы завышена
    в сто раз.
    """
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    raw = str(v)
    pct = "%" in raw
    s = raw.strip().replace("\u00a0", "").replace(" ", "").replace("%", "")
    if "," in s and "." in s:
        s = s.replace(",", "")          # 1,234.5
    else:
        s = s.replace(",", ".")         # 1234,5
    try:
        num = float(s)
    except ValueError:
        return 0.0
    return num / 100.0 if pct else num


_DATE_PATTERNS = ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%Y/%m/%d", "%d-%m-%Y")


def _day(v):
    """Дата в 'YYYY-MM-DD'. Не разобралась — пустая строка."""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if hasattr(v, "isoformat") and not isinstance(v, str):
        try:
            return v.isoformat()[:10]
        except Exception:
            return ""
    s = str(v).strip()
    m = re.search(r"(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    for fmt in _DATE_PATTERNS:
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return ""


_NAME_DATE = (
    re.compile(r"(20\d{2})[-_.](\d{2})[-_.](\d{2})"),      # 2026-08-08
    re.compile(r"(?<!\d)(\d{2})[-_.](\d{2})[-_.](20\d{2})"),  # 08.08.2026
    re.compile(r"(?<!\d)(20\d{2})(\d{2})(\d{2})(?!\d)"),      # 20260808
)


def day_from_name(name):
    """
    Дата из имени файла: '2026-08-08.xlsx', '08.08.2026.xlsx', '20260808.xlsx'.

    Нужна потому, что конструктор отчётов в кабинете НЕ умеет разрез по дням:
    в настройке группировок есть только товары, бренды и категории. Значит
    один файл может описывать только один день, и день этот приходится
    брать из имени — больше его в файле негде взять.
    """
    base = os.path.basename(str(name or ""))
    for i, rx in enumerate(_NAME_DATE):
        m = rx.search(base)
        if not m:
            continue
        a, b, c = m.groups()
        y, mo, d = (a, b, c) if len(a) == 4 else (c, b, a)
        try:
            return datetime(int(y), int(mo), int(d)).date().isoformat()
        except ValueError:
            continue
    return ""


def _rows_from_csv(data):
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4000]
    delim = ";" if sample.count(";") >= sample.count(",") else ","
    return [r for r in csv.reader(io.StringIO(text, newline=""), delimiter=delim)]


# Выгрузка заказов у заказчика приходит листом на миллион строк, из которых
# заполнено восемь тысяч: Excel растягивает лист до предела. Читать хвост
# бессмысленно и дорого, поэтому останавливаемся после длинной череды пустых.
EMPTY_TAIL = 200


# Признаки нужного листа, от самого надёжного к слабому. «Кластер доставки»
# встречается только в выгрузке заказов, «конверсия в корзину» — только в
# аналитической. Заказчик прислал рабочую книгу на десять листов, где первым
# лежал совсем другой лист, поэтому лист выбирается по шапке, а не по номеру.
SHEET_MARKERS = (
    ("кластер доставки",),
    ("конверсия в корзину", "уникальные посетители"),
    ("в корзину", "показы"),
)


def _pick_sheet(wb):
    """Лист с нужными данными. Не нашли по приметам — берём первый."""
    heads = {}
    for name in wb.sheetnames:
        cells = []
        for i, row in enumerate(wb[name].iter_rows(values_only=True)):
            if i > 4:
                break
            cells += [_norm(c) for c in (row or ()) if c]
        heads[name] = cells
    for markers in SHEET_MARKERS:
        for name in wb.sheetnames:
            if any(any(m in h for h in heads[name]) for m in markers):
                return wb[name]
    return wb[wb.sheetnames[0]]


def _rows_from_xlsx(data):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = _pick_sheet(wb)
    out, empty = [], 0
    for row in ws.iter_rows(values_only=True):
        if row is None or all(c is None or c == "" for c in row):
            empty += 1
            if empty >= EMPTY_TAIL:
                break
            out.append(list(row) if row else [])
            continue
        empty = 0
        out.append(list(row))
    while out and all(c is None or c == "" for c in out[-1]):
        out.pop()
    return out


def _match_by(header, spec, order):
    """Как _match_columns, но по произвольному словарю ролей."""
    norm = [_norm(h) for h in header]
    found, used = {}, set()
    for role in order:
        names = spec[role]
        for i, h in enumerate(norm):
            if i not in used and h in names:
                found[role], _ = i, used.add(i)
                break
        if role in found:
            continue
        for i, h in enumerate(norm):
            if i not in used and any(n in h for n in names):
                found[role], _ = i, used.add(i)
                break
    return found


def looks_like_orders(rows):
    """Выгрузка заказов? Узнаём по колонке «Кластер доставки»."""
    for row in rows[:15]:
        if row and any("кластер доставки" == _norm(c) for c in row):
            return True
    return False


def parse_orders(rows, source=""):
    """
    Выгрузка заказов -> {артикул: {кластер доставки: {день: штук}}}.

    Отменённые пропускаются. День берётся из «Принят в обработку» — именно
    по нему заказчик считает свою неделю.
    """
    head, cols = -1, {}
    for i, row in enumerate(rows[:15]):
        if not row:
            continue
        c = _match_by(row, ORDER_COLUMNS, ORDER_ORDER)
        if "cluster" in c and "day" in c and ("offer_id" in c or "sku" in c):
            head, cols = i, c
            break
    if head < 0:
        raise CabinetImportError(
            f"{source}: похоже на выгрузку заказов, но не нашёл разом "
            f"«Кластер доставки», «Принят в обработку» и «Артикул»")

    out, skipped = {}, 0
    for row in rows[head + 1:]:
        if not row:
            continue

        def cell(role):
            i = cols.get(role)
            return row[i] if i is not None and i < len(row) else None

        st = _norm(cell("status"))
        if st and any(w in st for w in CANCELLED):
            skipped += 1
            continue
        key = str(cell("offer_id") or "").strip() or str(cell("sku") or "").strip()
        cluster = str(cell("cluster") or "").strip()
        day = _day(cell("day"))
        if not key or not cluster or not day:
            continue
        qty = _num(cell("qty")) if "qty" in cols else 1.0
        out.setdefault(key, {}).setdefault(cluster, {})
        out[key][cluster][day] = out[key][cluster].get(day, 0.0) + (qty or 1.0)

    if not out:
        raise CabinetImportError(f"{source}: заказы разобраны, но строк нет")
    if skipped:
        log.info("%s: пропущено отменённых строк: %d", source, skipped)
    return out


def _find_header(rows, default_day=""):
    """
    Шапка не всегда в первой строке: выгрузки OZON любят пару строк
    с названием отчёта и периодом сверху. Ищем первую строку, где нашлась
    хоть одна нужная величина, а вместе с ней дата — либо колонкой, либо
    заранее известным днём из имени файла.
    """
    for i, row in enumerate(rows[:15]):
        if not row:
            continue
        cols = _match_columns(row)
        if not any(k in cols for k in VALUE_KEYS):
            continue
        if "day" in cols or default_day:
            return i, cols
    return -1, {}


def parse_rows(rows, source="", default_day=""):
    """
    Строки файла -> {ключ_товара: {день: {views, sessions, tocart, position}}}.
    Ключ — артикул, если он есть в файле, иначе sku.

    default_day — день, к которому отнести весь файл, когда колонки с датой
    в нём нет. Берётся из имени файла вызывающим кодом.
    """
    head, cols = _find_header(rows, default_day)
    if head < 0:
        sample = _norm(" | ".join(str(c) for c in (rows[0] if rows else [])))
        if any("день" not in _norm(str(c)) for c in (rows[0] if rows else [])) \
                and rows:
            raise CabinetImportError(
                f"{source}: не понял, за какой день файл. В кабинете разреза "
                f"по дням нет, поэтому дату надо взять из имени файла — "
                f"назовите его датой выгруженного дня, например "
                f"2026-08-08.xlsx. Первая строка файла: {sample[:200]}")
        raise CabinetImportError(f"{source}: не разобрал шапку файла")

    if "offer_id" not in cols and "sku" not in cols:
        raise CabinetImportError(
            f"{source}: в файле нет ни артикула, ни SKU — товары не с чем "
            f"сопоставить")

    out = {}
    bad_days = 0
    for row in rows[head + 1:]:
        if not row:
            continue

        def cell(role):
            i = cols.get(role)
            return row[i] if i is not None and i < len(row) else None

        day = _day(cell("day")) if "day" in cols else default_day
        if not day:
            bad_days += 1
            continue
        key = str(cell("offer_id") or "").strip() or str(cell("sku") or "").strip()
        if not key or key.lower() in ("none", "итого", "total"):
            continue

        rec = out.setdefault(key, {}).setdefault(
            day, {r: 0.0 for r in VALUE_KEYS})
        for role in VALUE_KEYS:
            if role in cols:
                v = _num(cell(role))
                if role in SUM_KEYS:
                    rec[role] += v
                elif v:
                    # позиция и конверсия — доли и места, их не складывают
                    rec[role] = v

    # Корзину кабинет часто отдаёт только процентом «конверсия в корзину из
    # карточки товара». Штуки из него получаются умножением на посетителей
    # карточки — это арифметика над двумя отданными числами, а не догадка.
    # Но считается она ТОЛЬКО по карточке: добавления из поиска и каталога
    # сюда не попадают, поэтому число выйдет ниже кабинетного «в корзину».
    if "tocart" not in cols and "conv_tocart" in cols and "sessions_pdp" in cols:
        for days in out.values():
            for rec in days.values():
                rec["tocart"] = rec["conv_tocart"] * rec["sessions_pdp"]
        log.info("%s: колонки «в корзину» нет — считаю из конверсии по "
                 "карточке, число будет ниже кабинетного", source)

    if bad_days:
        log.debug("%s: строк без разобранной даты: %d", source, bad_days)
    if not out:
        raise CabinetImportError(f"{source}: файл разобран, но данных в нём нет")
    return out


def parse_file(data, name=""):
    """
    Байты файла -> разобранные строки. Формат по расширению.

    Если колонки с датой в файле нет, день берётся из имени: кабинет не
    умеет разрез по дням, поэтому один файл = один день.
    """
    low = name.lower()
    if low.endswith(".csv") or low.endswith(".txt"):
        rows = _rows_from_csv(data)
    elif low.endswith(".xlsx") or low.endswith(".xlsm"):
        rows = _rows_from_xlsx(data)
    elif data[:2] == b"PK":
        rows = _rows_from_xlsx(data)
    else:
        rows = _rows_from_csv(data)
    if looks_like_orders(rows):
        return "orders", parse_orders(rows, source=name or "файл")
    return "metrics", parse_rows(rows, source=name or "файл",
                                 default_day=day_from_name(name))


# ------------------------------------------------------------------ источники

# Сколько файлов читать из папки магазина. Один файл — один день, за месяц
# их набирается три десятка; потолок нужен, чтобы случайно сваленная туда
# сотня файлов не растянула сбор.
MAX_FILES = 40


def _merge(dst, src):
    """Складывает разборы разных файлов. Один и тот же день перезаписывается."""
    for key, days in (src or {}).items():
        dst.setdefault(key, {}).update(days)
    return dst


def _merge_orders(dst, src):
    """Заказы из разных файлов: слияние на два уровня, артикул -> кластер."""
    for key, clusters in (src or {}).items():
        tgt = dst.setdefault(key, {})
        for cluster, days in clusters.items():
            tgt.setdefault(cluster, {}).update(days)
    return dst


def _empty():
    return {"metrics": {}, "orders": {}}


def _local_files(folder):
    if not os.path.isdir(folder):
        return []
    files = [os.path.join(folder, f) for f in os.listdir(folder)
             if f.lower().endswith((".xlsx", ".xlsm", ".csv", ".txt"))
             and not f.startswith("~$")]
    # свежие последними: если два файла про один день, победит новый
    return sorted(files, key=os.path.getmtime)[-MAX_FILES:]


def load_local(store_name, data_dir):
    """Все файлы из DATA_DIR/import/<Магазин>/ — метрики и заказы."""
    out, days, bad, n_ord = _empty(), set(), 0, 0
    for path in _local_files(os.path.join(data_dir, "import", store_name)):
        name = os.path.basename(path)
        try:
            with open(path, "rb") as f:
                kind, part = parse_file(f.read(), name)
        except CabinetImportError as e:
            bad += 1
            log.warning("[%s] %s", store_name, e)
            continue
        if kind == "orders":
            _merge_orders(out["orders"], part)
            n_ord += 1
        else:
            days |= {d for v in part.values() for d in v}
            _merge(out["metrics"], part)
    if out["metrics"]:
        log.info("[%s] выгрузка кабинета: дней %d, товаров %d",
                 store_name, len(days), len(out["metrics"]))
    if out["orders"]:
        log.info("[%s] выгрузка заказов: файлов %d, товаров %d "
                 "(продажи по кластеру доставки)",
                 store_name, n_ord, len(out["orders"]))
    if not out["metrics"] and not out["orders"] and bad:
        log.warning("[%s] ни один файл выгрузки прочитать не удалось", store_name)
    return out


def load_drive(store_name, folder_id, credentials_file):
    """
    Самый свежий файл из подпапки <Магазин> в папке Google Диска.

    Сервисный аккаунт тот же, что и для выгрузки отчётов, права на Диск у
    него уже есть — достаточно поделиться папкой с его почтой.
    """
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    creds = service_account.Credentials.from_service_account_file(
        credentials_file, scopes=["https://www.googleapis.com/auth/drive"])
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)

    q = (f"'{folder_id}' in parents and trashed = false and "
         f"mimeType = 'application/vnd.google-apps.folder'")
    subs = drive.files().list(q=q, fields="files(id,name)").execute().get("files", [])
    sub = next((s for s in subs if _norm(s["name"]) == _norm(store_name)), None)
    if not sub:
        log.info("[%s] на Диске нет подпапки с таким именем — пропускаю",
                 store_name)
        return {}

    q = f"'{sub['id']}' in parents and trashed = false"
    files = drive.files().list(
        q=q, orderBy="modifiedTime", pageSize=MAX_FILES,
        fields="files(id,name,mimeType,modifiedTime)").execute().get("files", [])
    files = [f for f in files if f["mimeType"] != "application/vnd.google-apps.folder"]
    if not files:
        log.info("[%s] подпапка на Диске пуста — пропускаю", store_name)
        return {}

    out, days, bad, n_ord = _empty(), set(), 0, 0
    for f in files:
        if f["mimeType"] == "application/vnd.google-apps.spreadsheet":
            data = drive.files().export(fileId=f["id"], mimeType="text/csv").execute()
            name = f["name"] + ".csv"
        else:
            data = drive.files().get_media(fileId=f["id"])
            data = data.execute() if hasattr(data, "execute") else data
            name = f["name"]
        try:
            kind, part = parse_file(data, name)
        except CabinetImportError as e:
            bad += 1
            log.warning("[%s] %s", store_name, e)
            continue
        if kind == "orders":
            _merge_orders(out["orders"], part)
            n_ord += 1
        else:
            days |= {d for v in part.values() for d in v}
            _merge(out["metrics"], part)

    if out["metrics"]:
        log.info("[%s] выгрузка кабинета с Диска: дней %d, товаров %d",
                 store_name, len(days), len(out["metrics"]))
    if out["orders"]:
        log.info("[%s] выгрузка заказов с Диска: файлов %d, товаров %d "
                 "(продажи по кластеру доставки)",
                 store_name, n_ord, len(out["orders"]))
    if not out["metrics"] and not out["orders"] and bad:
        log.warning("[%s] на Диске %d файлов, но ни один не прочитан", store_name, bad)
    return out


def load(store_name, cfg):
    """
    Данные кабинета для магазина. Сначала Диск, если он настроен, потом
    локальная папка. Любая ошибка — предупреждение, а не падение сбора:
    отсутствие выгрузки не должно ронять остальные четыре отчёта.
    """
    folder = getattr(cfg, "GOOGLE_IMPORT_FOLDER", "") or ""
    creds = getattr(cfg, "GOOGLE_CREDENTIALS_FILE", "") or ""
    if folder and creds and os.path.exists(creds):
        try:
            data = load_drive(store_name, folder, creds)
            if data["metrics"] or data["orders"]:
                return data
        except CabinetImportError as e:
            log.warning("[%s] выгрузку с Диска не разобрал: %s", store_name, e)
        except Exception as e:  # сеть, права, отозванный ключ
            log.warning("[%s] Google Диск недоступен (%s) — смотрю локально",
                        store_name, str(e)[:200])
    try:
        return load_local(store_name, getattr(cfg, "DATA_DIR", "data"))
    except CabinetImportError as e:
        log.warning("[%s] локальную выгрузку не разобрал: %s", store_name, e)
        return _empty()
