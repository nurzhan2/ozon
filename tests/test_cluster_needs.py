# -*- coding: utf-8 -*-
"""Потребность по кластерам, колонка ср/28 дней и шкала ДРР."""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("OZON_STORES",
                      '[{"name":"ТЕСТ","client_id":"1","api_key":"k"}]')
os.environ.setdefault("DATA_DIR", "/tmp/_clneeds")

from openpyxl import load_workbook

from ozon import reports as R
from ozon import excel as X
from ozon.seller_api import SellerAPI

ok = True


def check(label, cond, extra=""):
    global ok
    print(("  OK   " if cond else "  ПРОВАЛ ") + label + ("" if cond else f"  <- {extra}"))
    ok = ok and cond


print("\n1. Разбор ответа: берётся ads_cluster, а не ads")
row = SellerAPI._cluster_row({
    "offer_id": "ART-1", "sku": 111, "name": "Товар",
    "cluster_name": "Москва", "warehouse_name": "W1",
    "available_stock_count": 100, "requested_stock_count": 0,
    "transit_stock_count": 0,
    "ads": 549.3, "ads_cluster": 30.0,          # по магазину и по кластеру
    "idc": 2.1, "idc_cluster": 3.3,
    "days_without_sales_cluster": 4,
})
check("ads — кластерный", row["ads"] == 30.0, row["ads"])
check("общий сохранён отдельно", row["ads_all"] == 549.3, row.get("ads_all"))
check("idc — кластерный", row["idc"] == 3.3, row["idc"])
check("дни без продаж в кластере", row["no_sales_days"] == 4, row.get("no_sales_days"))

row_old = SellerAPI._cluster_row({
    "offer_id": "A", "cluster_name": "К", "ads": 12.0, "idc": 5.0,
    "available_stock_count": 1,
})
check("старый ответ без ads_cluster не ломается",
      row_old["ads"] == 12.0 and row_old["idc"] == 5.0, row_old)


print("\n2. Продажи недели раскладываются по доле кластера в ПРОДАЖАХ")
CLUSTERS = [
    # Москва продаёт втрое больше Урала, но лежит там почти весь остаток
    {"offer_id": "ART-1", "name": "Мазь", "cluster": "Москва", "warehouse": "W1",
     "available": 300, "requested": 0, "transit": 100, "ads": 30.0, "idc": 13.3},
    {"offer_id": "ART-1", "name": "Мазь", "cluster": "Урал", "warehouse": "W2",
     "available": 50, "requested": 0, "transit": 0, "ads": 10.0, "idc": 5.0},
    {"offer_id": "ART-1", "name": "Мазь", "cluster": "Минск", "warehouse": "W3",
     "available": 0, "requested": 0, "transit": 0, "ads": 0.0, "idc": 0.0},
]


class Store:
    name = "ТЕСТ"

    def __init__(self, rows=None, sales=280):
        self.rows = CLUSTERS if rows is None else rows
        self.sales = sales

    def cluster_stocks(self):
        return self.rows

    def products_for_period(self, a, b, only_in_stock=True, with_kpi=True):
        if self.sales is None:
            raise RuntimeError("аналитика недоступна")
        return {"ART-1": {"offer_id": "ART-1", "ordered_units": self.sales}}


class Cfg:
    TIMEZONE = "Europe/Moscow"
    DATA_DIR = "/tmp/_clneeds"
    OUTPUT_DIR = "/tmp/_clneeds/out"


TOTAL = "все кластеры"


def build(store):
    """Строки по кластерам; итоговая строка позиции пропускается."""
    ws = load_workbook(R.build_stocks([store], Cfg()))["ТЕСТ"]
    out = {}
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value and ws.cell(r, 2).value != TOTAL:
            out[ws.cell(r, 2).value] = {
                "sold7": ws.cell(r, 7).value,
                "ads28": ws.cell(r, 9).value,
                "need30": ws.cell(r, 10).value,
            }
    return ws, out


ws, res = build(Store())
check("Москва получила свою долю недели (30 из 40)",
      res["Москва"]["sold7"] == 210, res["Москва"])
check("Урал получил свою (10 из 40)", res["Урал"]["sold7"] == 70, res["Урал"])
check("кластер без продаж получил ноль", res["Минск"]["sold7"] == 0, res["Минск"])
check("сумма по кластерам = реальным продажам за неделю",
      sum(v["sold7"] for v in res.values()) == 280,
      sum(v["sold7"] for v in res.values()))
check("числа РАЗНЫЕ по кластерам — это и была жалоба заказчика",
      len({v["sold7"] for v in res.values()}) == 3, res)

print("\n3. Колонка «ср/28 дней» — ads_cluster как есть")
check("шапка на месте", ws.cell(1, 9).value == "ср/28 дней", ws.cell(1, 9).value)
check("Москва 30", res["Москва"]["ads28"] == 30.0, res["Москва"]["ads28"])
check("Урал 10", res["Урал"]["ads28"] == 10.0, res["Урал"]["ads28"])
check("порядок колонок не съехал",
      [ws.cell(1, c).value for c in (10, 11, 12)]
      == ["потреб 30д", "потреб 45д", "на сколько дней хватит остатков"],
      [ws.cell(1, c).value for c in (10, 11, 12)])

print("\n4. Потребность считается от своего кластера")
check("формула потребности смотрит на среднее и итог своей строки",
      res["Москва"]["need30"] == "=H3*30-F3", res["Москва"]["need30"])
check("зелёная подсветка закрытой потребности переехала на J:K",
      any(str(rng.sqref).startswith("J") for rng in ws.conditional_formatting),
      [str(rng.sqref) for rng in ws.conditional_formatting])

print("\n5. Запасные пути, когда данных меньше")
# нет продаж за неделю (аналитика отвалилась) — берём темп 28 дней
_, res2 = build(Store(sales=None))
check("без недельных продаж считаем ads_cluster x 7",
      res2["Москва"]["sold7"] == 210 and res2["Урал"]["sold7"] == 70, res2)

# OZON не дал продаж по кластерам вовсе — делим по остаткам
no_ads = [dict(c, ads=0.0) for c in CLUSTERS]
_, res3 = build(Store(rows=no_ads, sales=400))
# доли по остаткам: Москва 400/450, Урал 50/450 -> 356 и 44
check("без ads_cluster делим по остаткам, а не роняем отчёт",
      res3["Москва"]["sold7"] == 356 and res3["Урал"]["sold7"] == 44, res3)

print("\n6. ДРР: чем выше, тем краснее")
from openpyxl import Workbook

wb = Workbook()
w = wb.active
X.color_scale_inverted(w, "A1:A5")
rule = list(w.conditional_formatting)[0].rules[0]
check("минимум зелёный", rule.colorScale.color[0].rgb.endswith(X.SCALE_MAX),
      rule.colorScale.color[0].rgb)
check("максимум красный", rule.colorScale.color[-1].rgb.endswith(X.SCALE_MIN),
      rule.colorScale.color[-1].rgb)

w2 = wb.create_sheet("прямая")
X.color_scale(w2, "A1:A5")
rule2 = list(w2.conditional_formatting)[0].rules[0]
check("обычная шкала осталась прежней",
      rule2.colorScale.color[0].rgb.endswith(X.SCALE_MIN), rule2.colorScale.color[0].rgb)

print("\n7. Показы в динамике: нет данных — пусто, а не ноль")
rec = {"days": {"2026-08-09": {"revenue": 100, "ordered_units": 1,
                               "ad_spend": 10, "hits_view": 0}}}
v_no = R._agg(rec, ["2026-08-09"], has_views=False)
v_yes = R._agg(rec, ["2026-08-09"], has_views=True)
check("не посчитано — пусто", v_no["hits_view"] is None, v_no["hits_view"])
check("посчитано — число", v_yes["hits_view"] == 0, v_yes["hits_view"])
check("остальные метрики не тронуты", v_no["revenue"] == 100 and v_no["ad_spend"] == 10)

t = R._totals([("a", v_no), ("b", v_no)])
check("итог по пустым показам тоже пустой", t["hits_view"] is None, t["hits_view"])
t2 = R._totals([("a", v_no), ("b", v_yes)])
check("если хоть один день известен — итог число", t2["hits_view"] == 0, t2["hits_view"])
check("итоги денег складываются как прежде", t2["revenue"] == 200, t2["revenue"])

d = R._delta_rows([("a", v_yes)], [("a", v_no)])
check("разница показов без пары дней — пусто",
      d[0][1]["hits_view"] is None, d[0][1]["hits_view"])
check("разница по деньгам считается", d[0][1]["revenue"] == 0, d[0][1]["revenue"])

print("\n8. Один кластер — одна строка, склады складываются")
WAREHOUSES = []
for w, av, tr in [("W1", 0, 0), ("W2", 0, 1), ("W3", 1, 0), ("W4", 399, 0),
                  ("W5", 77, 0), ("W6", 0, 500), ("W7", 2, 0)]:
    WAREHOUSES.append({
        "offer_id": "Мазь VARICOSE",
        "name": "Мазь от варикоза вен на ногах, крем венотоник от отеков",
        "cluster": "Москва, МО и Дальние регионы", "warehouse": w,
        "available": av, "requested": 0, "transit": tr, "ads": 18.6, "idc": 5.0})
WAREHOUSES.append({
    "offer_id": "Мазь VARICOSE", "name": "Мазь от варикоза", "cluster": "Ростов",
    "warehouse": "W8", "available": 400, "requested": 0, "transit": 0,
    "ads": 8.7, "idc": 9.0})


class Store2(Store):
    def __init__(self):
        super().__init__(rows=WAREHOUSES, sales=191)

    def products_for_period(self, a, b, only_in_stock=True, with_kpi=True):
        return {"Мазь VARICOSE": {"offer_id": "Мазь VARICOSE",
                                  "ordered_units": self.sales}}


ws2 = load_workbook(R.build_stocks([Store2()], Cfg()))["ТЕСТ"]
rows2 = {ws2.cell(r, 2).value: r for r in range(2, ws2.max_row + 1)
         if ws2.cell(r, 2).value and ws2.cell(r, 2).value != TOTAL}
check("восемь складских строк свернулись в два кластера",
      len(rows2) == 2, sorted(rows2))
check("кластер назван один раз", set(rows2) == {"Москва, МО и Дальние регионы",
                                               "Ростов"}, set(rows2))
msk = rows2["Москва, МО и Дальние регионы"]
check("«доступно» просуммировано по складам",
      ws2.cell(msk, 3).value == 479, ws2.cell(msk, 3).value)
check("«в пути» просуммировано", ws2.cell(msk, 5).value == 501,
      ws2.cell(msk, 5).value)
check("ср/28 НЕ просуммирован — он и так на весь кластер",
      ws2.cell(msk, 9).value == 18.6, ws2.cell(msk, 9).value)
check("продажи недели поделены между двумя кластерами",
      ws2.cell(msk, 7).value == 130
      and ws2.cell(rows2["Ростов"], 7).value == 61,
      (ws2.cell(msk, 7).value, ws2.cell(rows2["Ростов"], 7).value))

print("\n9. В колонке «Артикул» — артикул, а не длинное наименование")
check("артикул", ws2.cell(msk, 1).value == "Мазь VARICOSE", ws2.cell(msk, 1).value)
check("длинное наименование не попало",
      "венотоник" not in str(ws2.cell(msk, 1).value), ws2.cell(msk, 1).value)

print("\n10. Расписание: у утреннего пакета есть минуты")
import importlib
import datetime as _dt
os.environ["MORNING_HOUR"] = "9"
os.environ["MORNING_MINUTE"] = "45"
os.environ["INTRADAY_HOURS"] = "12,16"
import worker
importlib.reload(worker)
check("минута прочитана", worker.MORNING_MINUTE == 45, worker.MORNING_MINUTE)
moment, kinds = worker.next_run_after(_dt.datetime(2026, 8, 11, 7, 0))
check("утренний пакет в 09:45",
      (moment.hour, moment.minute) == (9, 45), (moment.hour, moment.minute))
check("и это именно morning", kinds == ["morning"], kinds)
moment, kinds = worker.next_run_after(_dt.datetime(2026, 8, 11, 10, 0))
check("следующий — промежуточный в 12:00",
      (moment.hour, moment.minute, kinds) == (12, 0, ["intraday"]),
      (moment.hour, moment.minute, kinds))
moment, kinds = worker.next_run_after(_dt.datetime(2026, 8, 11, 20, 0))
check("после последнего часа переходим на утро следующего дня",
      (moment.day, moment.hour, moment.minute) == (12, 9, 45),
      (moment.day, moment.hour, moment.minute))

os.environ["MORNING_MINUTE"] = "0"
os.environ["MORNING_HOUR"] = "12"
importlib.reload(worker)
moment, kinds = worker.next_run_after(_dt.datetime(2026, 8, 11, 7, 0))
check("совпадение часа с промежуточным даёт оба вида",
      sorted(kinds) == ["intraday", "morning"], kinds)

print("\nИТОГ:", "все проверки пройдены" if ok else "ЕСТЬ ПРОВАЛЫ")
sys.exit(0 if ok else 1)
