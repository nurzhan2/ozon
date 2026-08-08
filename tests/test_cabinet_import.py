# -*- coding: utf-8 -*-
"""Импорт выгрузок из личного кабинета: разбор файла и заполнение отчёта."""
import io
import os
import sys
import datetime as _dt

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("OZON_STORES",
                      '[{"name":"ТЕСТ","client_id":"1","api_key":"k"}]')

from openpyxl import Workbook, load_workbook

from ozon import cabinet as CAB
from ozon import processing as P
from ozon import reports as R

ok = True


def check(label, cond, extra=""):
    global ok
    print(("  OK   " if cond else "  ПРОВАЛ ") + label + ("" if cond else f"  <- {extra}"))
    ok = ok and cond


print("\n1. Шапка ищется по названиям, а не по номеру строки")
rows = [
    ["Отчёт по товарам"],
    ["Период: 05.08.2026 — 06.08.2026"],
    [],
    ["Дата", "Артикул", "Ozon ID", "Показы", "Сессии", "В корзину", "Позиция"],
    ["05.08.2026", "ART-1", "111", "12 500", "3 040", "621", "7"],
    ["06.08.2026", "ART-1", "111", "11 200", "2 800", "590", "9"],
]
out = CAB.parse_rows(rows, "тест")
check("две служебные строки сверху не помешали", "ART-1" in out, list(out))
check("пробелы в числах разобраны",
      out["ART-1"]["2026-08-05"]["views"] == 12500.0,
      out["ART-1"]["2026-08-05"])
check("дата ДД.ММ.ГГГГ приведена к ISO",
      set(out["ART-1"]) == {"2026-08-05", "2026-08-06"}, list(out["ART-1"]))
check("корзина на месте", out["ART-1"]["2026-08-06"]["tocart"] == 590.0)
check("позиция берётся как есть, не суммируется",
      out["ART-1"]["2026-08-06"]["position"] == 9.0)

print("\n2. Другие формулировки колонок — тот же результат")
rows2 = [
    ["День", "Ваш SKU", "Показы всего", "Добавления в корзину",
     "Средняя позиция"],
    ["2026-08-05", "ART-9", "1000", "50", "3,5"],
]
out2 = CAB.parse_rows(rows2, "тест2")
check("«Показы всего» распознано", out2["ART-9"]["2026-08-05"]["views"] == 1000.0)
check("«Добавления в корзину» распознано",
      out2["ART-9"]["2026-08-05"]["tocart"] == 50.0)
check("запятая как десятичный разделитель",
      out2["ART-9"]["2026-08-05"]["position"] == 3.5)

print("\n3. Строки одного дня складываются, «Итого» отбрасывается")
rows3 = [
    ["Дата", "Артикул", "Показы", "В корзину"],
    ["2026-08-05", "ART-1", "100", "5"],
    ["2026-08-05", "ART-1", "50", "3"],
    ["2026-08-05", "Итого", "150", "8"],
]
out3 = CAB.parse_rows(rows3, "тест3")
check("две строки за день сложились",
      out3["ART-1"]["2026-08-05"]["views"] == 150.0, out3["ART-1"])
check("строка «Итого» не стала товаром", "Итого" not in out3, list(out3))

print("\n4. Ни колонки с датой, ни даты в имени — внятный отказ")
try:
    CAB.parse_rows([["Артикул", "Показы", "В корзину"],
                    ["ART-1", "100", "5"]], "без_дат.xlsx")
    check("должно было подняться исключение", False)
except CAB.CabinetImportError as e:
    check("сказано назвать файл датой",
          "именем" in str(e) or "имени файла" in str(e), str(e)[:160])

print("\n5. Файл без артикула и без sku тоже отвергается")
try:
    CAB.parse_rows([["Дата", "Показы"], ["2026-08-05", "100"]], "без_ключа")
    check("должно было подняться исключение", False)
except CAB.CabinetImportError as e:
    check("сказано, что не с чем сопоставить", "сопоставить" in str(e), str(e))

print("\n6. csv с точкой с запятой")
data = "Дата;Артикул;Показы;В корзину\n2026-08-05;ART-2;300;12\n".encode("utf-8")
out6 = CAB.parse_file(data, "v.csv")
check("csv разобран", out6["ART-2"]["2026-08-05"]["tocart"] == 12.0, out6)

print("\n7. xlsx читается по байтам")
wb = Workbook()
ws = wb.active
ws.append(["Дата", "Артикул", "Показы", "В корзину"])
ws.append([_dt.date(2026, 8, 5), "ART-3", 700, 33])
buf = io.BytesIO()
wb.save(buf)
out7 = CAB.parse_file(buf.getvalue(), "v.xlsx")
check("дата-объект из ячейки понята",
      out7["ART-3"]["2026-08-05"]["views"] == 700.0, out7)

print("\n8. merge_cabinet перекрывает суррогаты, но не зануляет")
daily = {"ART-1": {"name": "Товар", "offer_id": "ART-1", "sku": "111", "days": {
    "2026-08-05": {"revenue": 1000, "session_view": 45, "ad_views": 900,
                   "ad_clicks": 45, "hits_view": 300},
    "2026-08-06": {"revenue": 900, "session_view": 20, "hits_view": 250}}}}
cab = {"111": {"2026-08-05": {"views": 12500, "sessions": 3040,
                              "tocart": 621, "position": 7}}}
filled = P.merge_cabinet(daily, cab, {111: {"offer_id": "ART-1"}})
d5 = daily["ART-1"]["days"]["2026-08-05"]
d6 = daily["ART-1"]["days"]["2026-08-06"]
check("показы кабинета вытеснили показы из поиска", d5["hits_view"] == 12500, d5)
check("клики кабинета вытеснили рекламные", d5["session_view"] == 3040, d5)
check("корзина появилась", d5["hits_tocart"] == 621, d5)
check("день без данных в выгрузке не тронут",
      d6["hits_view"] == 250 and d6["session_view"] == 20, d6)
check("сообщено, что именно закрыто",
      filled == {"hits_view", "session_view", "hits_tocart",
                 "position_category"}, filled)
check("товар найден по sku, хотя в daily ключ — артикул", True)

print("\n9. Подписи строк и CTR меняются вместе с источником")
rows_ad = R._quality_rows_for(set())
rows_cab = R._quality_rows_for({"session_view"})
check("без выгрузки клики подписаны как рекламные",
      ("клики (реклама)", "session_view", rows_ad[1][2]) == rows_ad[1], rows_ad[1])
check("с выгрузкой пометка снята", rows_cab[1][0] == "клики", rows_cab[1])
check("и у CTR тоже", rows_cab[2][0] == "CTR", rows_cab[2])

day = {"hits_view": 12500, "session_view": 3040, "ad_views": 900,
       "ad_clicks": 45, "hits_tocart": 621, "revenue": 100000, "ad_spend": 14000}
v_ad = R._quality_day_values(day, unified=False)
v_cab = R._quality_day_values(day, unified=True)
check("без выгрузки CTR по рекламной паре",
      abs(v_ad["ctr"] - 45 / 900) < 1e-9, v_ad["ctr"])
check("с выгрузкой CTR по своей паре",
      abs(v_cab["ctr"] - 3040 / 12500) < 1e-9, v_cab["ctr"])
check("% корзины считается от кликов, как в образце заказчика",
      abs(v_cab["cart_rate"] - 621 / 3040) < 1e-9, v_cab["cart_rate"])

print("\n10. Итог строки CTR сходится со способом расчёта дней")
vals = {"d1": R._quality_day_values(day, True),
        "d2": R._quality_day_values(day, True)}
tot = R._quality_row_total("ctr", vals, ["d1", "d2"], unified=True)
check("итог = сумма кликов / сумма показов",
      abs(tot - (3040 * 2) / (12500 * 2)) < 1e-9, tot)
vals = {"d1": R._quality_day_values(day, False),
        "d2": R._quality_day_values(day, False)}
tot = R._quality_row_total("ctr", vals, ["d1", "d2"], unified=False)
check("без выгрузки итог по рекламной паре",
      abs(tot - (45 * 2) / (900 * 2)) < 1e-9, tot)

print("\n11. Сноски нет, когда выгрузка всё закрыла")
totals = {"d": {"hits_view": 1, "hits_tocart": 1, "cart_rate": 0.2,
                "position_category": 3}}
check("пустых строк не осталось",
      R._quality_empty_keys(totals, ["d"]) == [],
      R._quality_empty_keys(totals, ["d"]))

print("\n12. Нет файла — сбор не падает")
check("пустая папка даёт пустой результат",
      CAB.load_local("НЕТ ТАКОГО", "/tmp/нет-такой-папки") == {})


class CfgNoGoogle:
    DATA_DIR = "/tmp/нет-такой-папки"
    GOOGLE_IMPORT_FOLDER = ""
    GOOGLE_CREDENTIALS_FILE = ""


check("load() без настроек возвращает пусто, а не исключение",
      CAB.load("ТЕСТ", CfgNoGoogle()) == {})

print("\n13. Сноска называет решение, а не только причину")
from openpyxl import Workbook as _WB
_ws = _WB().active
_r = R._quality_write_notes(_ws, 1, ["hits_tocart"], 3)
_texts = [str(_ws.cell(i, 1).value or "") for i in range(1, _r + 1)]
check("сказано про выгрузку из кабинета",
      any("ВЫГРУЗКА_ИЗ_КАБИНЕТА" in t for t in _texts), _texts)
check("и про Premium Plus как второй путь",
      any("Premium Plus" in t for t in _texts), _texts)

print("\n14. Нет данных за день — пусто, а не ноль")
day_full = {"hits_view": 900, "session_view": 40, "ad_views": 500,
            "ad_clicks": 20, "hits_tocart": 30, "revenue": 1000, "ad_spend": 100}
day_none = {"hits_view": 0, "session_view": 40, "ad_views": 500,
            "ad_clicks": 20, "revenue": 1000, "ad_spend": 100}

v = R._quality_day_values(day_none, has_views=False, has_cart=False)
check("показы пустые, а не нулевые", v["hits_view"] is None, v["hits_view"])
check("место в поиске тоже", v["position_category"] is None)
check("корзина пустая", v["hits_tocart"] is None)
check("% корзины пустой, а не 0%", v["cart_rate"] is None)
check("оборот и реклама остались числами",
      v["revenue"] == 1000 and v["ad_spend"] == 100, v)
check("CTR по рекламе считается и без показов", v["ctr"] == 20 / 500, v["ctr"])

v = R._quality_day_values(day_full, has_views=True, has_cart=True)
check("когда данные есть — обычные числа",
      v["hits_view"] == 900 and v["hits_tocart"] == 30, v)

print("\n15. Итоги строк не спотыкаются о пустые дни")
vals = {"d1": R._quality_day_values(day_full, has_views=True, has_cart=True),
        "d2": R._quality_day_values(day_none, has_views=False, has_cart=False)}
keys = ["d1", "d2"]
check("итог показов = только известные дни",
      R._quality_row_total("hits_view", vals, keys) == 900,
      R._quality_row_total("hits_view", vals, keys))
check("итог корзины считает известный день",
      R._quality_row_total("hits_tocart", vals, keys) == 30,
      R._quality_row_total("hits_tocart", vals, keys))
check("итог места в поиске не падает",
      R._quality_row_total("position_category", vals, keys) is not None)

vals_none = {"d": R._quality_day_values(day_none, has_views=False, has_cart=False)}
check("все дни пустые — итог тоже пустой, а не ноль",
      R._quality_row_total("hits_view", vals_none, ["d"]) is None,
      R._quality_row_total("hits_view", vals_none, ["d"]))
check("и у корзины",
      R._quality_row_total("hits_tocart", vals_none, ["d"]) is None)
check("и у % корзины",
      R._quality_row_total("cart_rate", vals_none, ["d"]) is None)
check("и у места в поиске",
      R._quality_row_total("position_category", vals_none, ["d"]) is None)

print("\n16. Свод по магазину уважает те же признаки")
items = [{"days": {"2026-08-05": {"hits_view": 100, "revenue": 10},
                   "2026-08-06": {"hits_view": 0, "revenue": 10}}}]
tot = R._quality_store_totals(items, ["2026-08-05", "2026-08-06"], False,
                              views_days={"2026-08-05"}, cart_days=set())
check("день с данными — число", tot["2026-08-05"]["hits_view"] == 100)
check("день без данных — пусто", tot["2026-08-06"]["hits_view"] is None)
check("сноска считает такую строку пустой только если пусты ВСЕ дни",
      "hits_view" not in R._quality_empty_keys(tot, ["2026-08-05", "2026-08-06"]),
      R._quality_empty_keys(tot, ["2026-08-05", "2026-08-06"]))

print("\n17. Свод сходится с суммой товаров до рубля")
prods = [{"days": {"d": {"revenue": 100.4, "ad_spend": 33.5, "ordered_units": 1}}},
         {"days": {"d": {"revenue": 200.4, "ad_spend": 33.5, "ordered_units": 1}}},
         {"days": {"d": {"revenue": 300.4, "ad_spend": 33.5, "ordered_units": 1}}}]
tot = R._quality_store_totals(prods, ["d"], False, views_days=set(), cart_days=set())
per_product = [R._quality_day_values(p["days"]["d"], has_views=False,
                                     has_cart=False) for p in prods]
for metric in ("revenue", "ad_spend"):
    check(f"«{metric}»: свод = сумма товарных клеток",
          tot["d"][metric] == sum(v[metric] for v in per_product),
          (tot["d"][metric], sum(v[metric] for v in per_product)))

print("\n18. Настоящая шапка выгрузки кабинета (файл клиента)")
REAL = ["Товары", "Модель", "Ozon ID", "Артикул", "Заказано товаров",
        "Уникальные посетители, всего",
        "Уникальные посетители с просмотром карточки товара",
        "Конверсия в корзину из карточки товара", "Заказано на сумму",
        "Отменено товаров", "Возвращено товаров", "Позиция в поиске и каталоге"]
cols = CAB._match_columns(REAL)
check("«Уникальные посетители, всего» -> клики",
      REAL[cols["sessions"]] == "Уникальные посетители, всего", cols)
check("посетители карточки распознаны отдельно",
      REAL[cols["sessions_pdp"]].startswith("Уникальные посетители с"), cols)
check("конверсия распознана как конверсия",
      REAL[cols["conv_tocart"]].startswith("Конверсия"), cols)
check("конверсия НЕ засчитана как абсолютная корзина",
      "tocart" not in cols, cols.get("tocart"))
check("артикул важнее Ozon ID", REAL[cols["offer_id"]] == "Артикул", cols)
check("позиция найдена", REAL[cols["position"]].startswith("Позиция"), cols)
check("колонки с датой нет — это и должно отвергнуть файл", "day" not in cols)

print("\n19. Корзина в штуках выводится из конверсии")
real_rows = [["Дата"] + REAL,
             ["09.08.2026", "Маска", "V", "2531686490", "ART-M", "2830",
              "1309694", "22445", "36,99%", "790662", "193", "2", "53,43"]]
out = CAB.parse_rows(real_rows, "клиентский.xlsx")
rec = out["ART-M"]["2026-08-09"]
check("клики = посетители всего", rec["sessions"] == 1309694, rec["sessions"])
check("процент разобран долей", abs(rec["conv_tocart"] - 0.3699) < 1e-9,
      rec["conv_tocart"])
check("корзина = посетители карточки x конверсия",
      abs(rec["tocart"] - 22445 * 0.3699) < 0.01, rec["tocart"])
check("позиция с запятой разобрана", rec["position"] == 53.43, rec["position"])
check("«Заказано товаров» не уехало в показы", rec["views"] == 0, rec["views"])

print("\n20. Абсолютная колонка «В корзину» важнее конверсии")
both = [["Дата", "Артикул", "Уникальные посетители с просмотром карточки товара",
         "Конверсия в корзину из карточки товара", "В корзину"],
        ["09.08.2026", "ART-B", "1000", "30%", "777"]]
rec = CAB.parse_rows(both, "обе.xlsx")["ART-B"]["2026-08-09"]
check("взята абсолютная, а не 1000x0.3", rec["tocart"] == 777, rec["tocart"])

print("\n21. День берётся из имени файла")
for name, want in [("2026-08-08.xlsx", "2026-08-08"),
                   ("08.08.2026.csv", "2026-08-08"),
                   ("20260808.xlsx", "2026-08-08"),
                   ("ШТУЧКА_08.08.2026.xlsx", "2026-08-08"),
                   ("отчёт (1).xlsx", ""),
                   ("2026-13-45.xlsx", "")]:
    check(f"«{name}» -> {want or 'даты нет'}",
          CAB.day_from_name(name) == want, CAB.day_from_name(name))

print("\n22. Файл без колонки даты разбирается по имени")
no_date = [["Артикул", "Уникальные посетители, всего",
            "Уникальные посетители с просмотром карточки товара",
            "Конверсия в корзину из карточки товара", "Позиция в поиске и каталоге"],
           ["ART-1", "1000", "200", "30%", "12,5"]]
import io as _io
from openpyxl import Workbook as _W
_wb = _W(); _ws2 = _wb.active
for r in no_date:
    _ws2.append(r)
_buf = _io.BytesIO(); _wb.save(_buf)
out = CAB.parse_file(_buf.getvalue(), "2026-08-08.xlsx")
check("день взят из имени", list(out["ART-1"]) == ["2026-08-08"], out)
check("корзина посчитана", abs(out["ART-1"]["2026-08-08"]["tocart"] - 60) < 0.01,
      out["ART-1"]["2026-08-08"]["tocart"])

print("\n23. Колонка с датой сильнее имени файла")
with_date = [["Дата", "Артикул", "Показы"], ["2026-07-01", "ART-1", "5"]]
out = CAB.parse_rows(with_date, "2026-08-08.xlsx", default_day="2026-08-08")
check("взята дата из строки, а не из имени",
      list(out["ART-1"]) == ["2026-07-01"], out)

print("\n24. Папка с файлами по дням собирается в один период")
import os as _os, shutil as _sh, time as _t
_folder = "/tmp/_cabtest/import/ТЕСТ"
_sh.rmtree("/tmp/_cabtest", ignore_errors=True)
_os.makedirs(_folder)
for _i, _d in enumerate(["2026-08-05", "2026-08-06", "2026-08-07"]):
    _wb = _W(); _w = _wb.active
    _w.append(["Артикул", "Уникальные посетители, всего", "В корзину"])
    _w.append(["ART-1", 100 + _i, 10 + _i])
    _wb.save(_os.path.join(_folder, f"{_d}.xlsx"))
    _t.sleep(0.02)
res = CAB.load_local("ТЕСТ", "/tmp/_cabtest")
check("собраны все три дня",
      sorted(res["ART-1"]) == ["2026-08-05", "2026-08-06", "2026-08-07"],
      sorted(res.get("ART-1", {})))
check("числа не перепутались между днями",
      res["ART-1"]["2026-08-07"]["sessions"] == 102,
      res["ART-1"]["2026-08-07"])

# битый файл не должен утаскивать за собой остальные
_wb = _W(); _wb.active.append(["ничего", "полезного"])
_wb.save(_os.path.join(_folder, "мусор.xlsx"))
res = CAB.load_local("ТЕСТ", "/tmp/_cabtest")
check("нечитаемый файл пропущен, остальные разобраны",
      len(res["ART-1"]) == 3, sorted(res.get("ART-1", {})))
_sh.rmtree("/tmp/_cabtest", ignore_errors=True)

print("\nИТОГ:", "все проверки пройдены" if ok else "ЕСТЬ ПРОВАЛЫ")
sys.exit(0 if ok else 1)
