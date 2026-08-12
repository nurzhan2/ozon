# -*- coding: utf-8 -*-
"""
Во всех пяти отчётах в первой колонке артикул, а не наименование.

Заказчик: «тут так и остались длинные наименования». В отчёте 4 артикул
поставили раньше, в остальных четырёх осталось наименование по сто с лишним
символов. Здесь проверяется, что подпись товара везде одна и та же и что
переход не ломает сравнение со вчерашними снимками.
"""
import datetime as _dt
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("OZON_STORES",
                      '[{"name":"ТЕСТ","client_id":"1","api_key":"k"}]')
os.environ.setdefault("DATA_DIR", "/tmp/_lbltest")

from openpyxl import load_workbook

from ozon import dates as D
from ozon import reports as R
from ozon import snapshots as S

ok = True


def check(label, cond, extra=""):
    global ok
    print(("  OK   " if cond else "  ПРОВАЛ ") + label + ("" if cond else f"  <- {extra}"))
    ok = ok and cond


LONG = ("Мазь от варикоза вен на ногах, крем венотоник от отеков и тяжести, "
        "гель для ног от усталости 50 мл")
ART = "Мазь VARICOSE"
LONG2 = "Тушь для ресниц Sky High Суперобъем - Эффект накладных ресниц, чёрная"
ART2 = "Тушь SKY-HIGH"

TZ = "Europe/Moscow"
TODAY = D.now_tz(TZ).date()
YDAY = D.yesterday(TZ)
D_TODAY, D_YDAY = D.d(TODAY), D.d(YDAY)
MONTH_DAYS = [D.d(x) for x in R._daterange(D.month_start(ref=YDAY, tz_name=TZ), YDAY)]


def day(rev, units, spend=0.0, views=0.0):
    return {"revenue": rev, "ordered_units": units, "ad_spend": spend,
            "hits_view": views, "session_view": 0, "hits_tocart": 0,
            "cancellations": 0, "position_category": 0}


def rec(offer, name, days):
    return {"offer_id": offer, "name": name, "sku": "111", "days": days}


class Store:
    """Сборщик с двумя товарами: у обоих длинное имя и короткий артикул."""
    name = "ТЕСТ"

    def __init__(self):
        self.days_with_views = set(MONTH_DAYS) | {D_TODAY, D_YDAY}
        self.days_with_cart = set()
        self.cabinet_filled = set()

    def daily_by_product(self, date_from, date_to, only_in_stock=True):
        keys = [D.d(x) for x in R._daterange(date_from, date_to)]
        d1 = {k: day(1000, 4, 100, 500) for k in keys}
        d2 = {k: day(700, 3, 50, 300) for k in keys}
        return {ART: rec(ART, LONG, d1), ART2: rec(ART2, LONG2, d2)}, {}


class Cfg:
    TIMEZONE = TZ
    DATA_DIR = "/tmp/_lbltest"
    OUTPUT_DIR = "/tmp/_lbltest/out"
    SNAPSHOTS_DIR = "/tmp/_lbltest/snap"
    CUMULATIVE_DAYS = 0


def col_a(path, sheet=None):
    wb = load_workbook(path)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)
            if ws.cell(r, 1).value]


def clean():
    shutil.rmtree("/tmp/_lbltest", ignore_errors=True)


clean()

print("\n1. Хелпер подписи")
check("артикул сильнее наименования",
      R._label({"offer_id": ART, "name": LONG}) == ART,
      R._label({"offer_id": ART, "name": LONG}))
check("без артикула остаётся наименование",
      R._label({"offer_id": "", "name": LONG}) == LONG)
check("нет ни того, ни другого — пустая строка, а не None",
      R._label({}) == "", repr(R._label({})))

print("\n2. Отчёт 1 «Общая сводная»")
a = col_a(R.build_cumulative_sales([Store()], Cfg()))
check("артикулы на месте", ART in a and ART2 in a, a[:4])
check("длинных наименований нет", LONG not in a and LONG2 not in a, a[:4])

print("\n3. Отчёт 2 «Динамика день ко дню»")
a = col_a(R.build_day_over_day([Store()], Cfg()), "ТЕСТ")
check("артикулы на месте", ART in a and ART2 in a, a[:6])
check("длинных наименований нет", LONG not in a, a[:6])
check("строка «Итог» на месте", "Итог" in a, a)

print("\n4. Отчёт 3 «Качественные показатели»")
a = col_a(R.build_quality([Store()], Cfg()), "ТЕСТ")
check("артикулы в заголовках блоков", ART in a and ART2 in a, a[:6])
check("длинных наименований нет", LONG not in a and LONG2 not in a, a[:6])
check("сводный блок «по аналитике» не тронут", "по аналитике" in a, a[:3])

print("\n5. Отчёт 4 «Остатки» — как и было")
CL = [{"offer_id": ART, "name": LONG, "cluster": "Москва", "warehouse": "W1",
       "available": 100, "requested": 0, "transit": 0, "ads": 10.0, "idc": 5.0}]


class StockStore(Store):
    def cluster_stocks(self):
        return CL

    def products_for_period(self, a, b, only_in_stock=True, with_kpi=True):
        return {ART: {"offer_id": ART, "ordered_units": 70}}


a = col_a(R.build_stocks([StockStore()], Cfg()), "ТЕСТ")
check("артикул", ART in a, a)
check("наименования нет", LONG not in a, a)

print("\n6. Отчёт 5 «Промежуточный» — тот, на который пожаловались")
a = col_a(R.build_intraday([Store()], Cfg()), "ТЕСТ")
check("артикулы на месте", ART in a and ART2 in a, a[:6])
check("длинных наименований нет", LONG not in a and LONG2 not in a, a[:6])

print("\n7. Вчерашние снимки со старыми подписями не раздваивают товары")
clean()
hour = D.now_tz(TZ).hour
# снимок вчерашнего дня, сделанный прежней версией — ключи наименованиями
old_snap = {LONG: {"revenue": 900, "ordered_units": 3, "ad_spend": 90,
                   "drr": 0.1, "hits_view": 400},
            LONG2: {"revenue": 600, "ordered_units": 2, "ad_spend": 40,
                    "drr": 0.07, "hits_view": 200}}
S.save(Cfg.SNAPSHOTS_DIR, "ТЕСТ", D_YDAY, hour, old_snap)
path = R.build_intraday([Store()], Cfg())
a = col_a(path, "ТЕСТ")
check("длинные наименования из старого снимка не всплыли",
      LONG not in a and LONG2 not in a, [x for x in a if len(str(x)) > 40])
check("товары не раздвоились: два артикула в каждом из трёх блоков",
      a.count(ART) == 3 and a.count(ART2) == 3, (a.count(ART), a.count(ART2)))
ws = load_workbook(path)["ТЕСТ"]
titles = [str(ws.cell(r, 1).value) for r in range(1, ws.max_row + 1)
          if str(ws.cell(r, 1).value or "").startswith("вчера")]
check("в подписи блока сказано, что сравнение с полным вчерашним днём",
      titles and "снимка за вчера" in titles[0], titles)

print("\n8. Сегодняшний снимок пишется уже артикулами")
saved = S.load(Cfg.SNAPSHOTS_DIR, "ТЕСТ", D_TODAY, hour)
check("ключи снимка — артикулы", sorted(saved or {}) == sorted([ART, ART2]),
      sorted(saved or {}))

print("\n9. Завтра сравнение со свежим снимком работает как прежде")
# тот же снимок, но уже с новыми подписями — ветка-заглушка НЕ должна сработать
clean()
new_snap = {ART: {"revenue": 900, "ordered_units": 3, "ad_spend": 90,
                  "drr": 0.1, "hits_view": 400}}
S.save(Cfg.SNAPSHOTS_DIR, "ТЕСТ", D_YDAY, hour, new_snap)
path = R.build_intraday([Store()], Cfg())
ws = load_workbook(path)["ТЕСТ"]
titles = [str(ws.cell(r, 1).value) for r in range(1, ws.max_row + 1)
          if str(ws.cell(r, 1).value or "").startswith("вчера")]
check("снимок принят, оговорки в подписи нет",
      titles and "снимка за вчера" not in titles[0], titles)

clean()

print("\nИТОГ:", "все проверки пройдены" if ok else "ЕСТЬ ПРОВАЛЫ")
sys.exit(0 if ok else 1)
