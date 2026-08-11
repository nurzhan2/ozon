# -*- coding: utf-8 -*-
"""
Пять отчётов OZON — макеты строго по образцам заказчика.

  1) cumulative  «Общая сводная по продажам» — один лист, магазины блоками,
                 колонки = дни (с 1-го числа по вчера), «Общий итог», «ср/день».
                 Образец: Отчет неделя_кол-во.xlsx
  2) dod         «Динамика день ко дню» — лист на магазин, три блока:
                 вчера / позавчера / динамика. Колонки: Сумма, Количество,
                 реклама, ДРР, Показы + строка «Итог».
                 Образец: динамика день ко дню (лист 3).csv
  3) quality     «Качественные показатели по каждому товару» — лист на магазин,
                 блок на товар: строки-метрики (показы, клики, CTR, корзина,
                 % корзины, купили без отмен, отмен, место в поиске, оборот,
                 реклама, ДРР %), колонки = дни. Накопительно, только на остатках.
  4) stocks      «По каждой позиции на остатках» — лист на магазин, строки
                 артикул × кластер: Доступно к продаже, В заявках на поставку,
                 В поставках в пути, Итог, прод 7д, среднее, потреб 30д,
                 потреб 45д, на сколько дней хватит остатков (формулы Excel).
                 Образец: Распред мазь 5 ядов Бьютифул.xlsx
  5) intraday    Тот же макет, что 2, но сегодня на время T к вчера на время T.

Во всех отчётах исключаются артикулы с меткой OUT; накопительные и
качественные строятся только по товарам, у которых есть остаток.
"""

import os
import logging
from datetime import timedelta

from openpyxl import Workbook

from . import excel as X
from . import dates as D
from . import snapshots as S
from . import processing as P

log = logging.getLogger("ozon.reports")


# ============================================================ утилиты

def _out(cfg, filename):
    os.makedirs(cfg.OUTPUT_DIR, exist_ok=True)
    return os.path.join(cfg.OUTPUT_DIR, filename)


def _new_wb():
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _daterange(date_from, date_to):
    days, cur = [], date_from
    while cur <= date_to:
        days.append(cur)
        cur += timedelta(days=1)
    return days


def _period(cfg):
    """Период накопительных отчётов: с 1-го числа по вчера (или последние N дней)."""
    tz = cfg.TIMEZONE
    date_to = D.yesterday(tz)
    n = getattr(cfg, "CUMULATIVE_DAYS", 0)
    if n and n > 0:
        return date_to - timedelta(days=n - 1), date_to
    return D.month_start(ref=date_to, tz_name=tz), date_to


def _safe_div(a, b):
    return (a / b) if b else 0.0


# ============================================================ 1. Общая сводная

def build_cumulative_sales(collectors, cfg):
    """Магазины блоками на одном листе, дни колонками, продажи в штуках."""
    tz = cfg.TIMEZONE
    date_from, date_to = _period(cfg)
    days = _daterange(date_from, date_to)
    day_keys = [D.d(x) for x in days]
    stamp = D.now_tz(tz).strftime("%Y-%m-%d")

    wb = _new_wb()
    ws = wb.create_sheet(X.safe_title(date_to.strftime("%d.%m")))

    r = 1
    for colr in collectors:
        daily, _ = colr.daily_by_product(date_from, date_to, only_in_stock=True)

        # строки: товар + продажи по дням
        rows = []
        for rec in daily.values():
            per_day = [int(round(rec["days"].get(k, {}).get("ordered_units", 0) or 0))
                       for k in day_keys]
            total = sum(per_day)
            if total <= 0:
                continue
            rows.append((rec["name"] or rec["offer_id"], per_day, total))
        rows.sort(key=lambda x: x[2], reverse=True)

        # --- шапка блока ---
        c = ws.cell(r, 1, value=colr.name)
        X.style_header_cell(c, yellow=True)
        for i, dt in enumerate(days):
            c = ws.cell(r, 2 + i, value=dt)
            X.style_header_cell(c)
            c.number_format = X.FMT_DATE
        c = ws.cell(r, 2 + len(days), value="Общий итог")
        X.style_header_cell(c)
        c = ws.cell(r, 3 + len(days), value="ср/день")
        X.style_header_cell(c)

        first_data_row = r + 1
        for name, per_day, total in rows:
            r += 1
            cc = ws.cell(r, 1, value=name)
            X.style_body_cell(cc)
            cc.alignment = X.LEFT
            for i, v in enumerate(per_day):
                X.style_body_cell(ws.cell(r, 2 + i, value=v), X.FMT_INT)
            X.style_body_cell(ws.cell(r, 2 + len(days), value=total), X.FMT_INT, bold=True)
            avg = round(total / len(days)) if days else 0
            X.style_body_cell(ws.cell(r, 3 + len(days), value=avg), X.FMT_INT, bold=True)

        # --- цветовые шкалы: отдельно на дни и на «Общий итог» (как в образце) ---
        if rows:
            X.color_scale(ws, f"B{first_data_row}:{X.col(1 + len(days))}{r}")
            X.color_scale(ws, f"{X.col(2 + len(days))}{first_data_row}:"
                              f"{X.col(2 + len(days))}{r}")
        r += 2   # пустая строка между магазинами

    X.set_widths(ws, [40] + [10] * len(days) + [13, 10])
    ws.freeze_panes = "B2"
    X.page_setup(ws)

    path = _out(cfg, f"01_svodnaya_prodazhi_{stamp}.xlsx")
    wb.save(path)
    log.info("отчёт 1 готов: %s", path)
    return path


# ============================================================ 2 и 5. Динамика

DOD_COLUMNS = [
    ("revenue", "Сумма", X.FMT_MONEY),
    ("ordered_units", "Количество", X.FMT_INT),
    ("ad_spend", "реклама", X.FMT_MONEY),
    ("drr", "ДРР", X.FMT_PCT),
    ("hits_view", "Показы", X.FMT_INT),
]


def _dod_block(ws, r, title, rows, totals):
    """
    Один блок таблицы: шапка (жёлтая ячейка с подписью периода) + товары + Итог.
    rows: список (название, {метрика: значение}); totals: {метрика: значение}.
    Возвращает номер следующей свободной строки.
    """
    c = ws.cell(r, 1, value=title)
    X.style_header_cell(c, yellow=True)
    for i, (_, header, _) in enumerate(DOD_COLUMNS):
        X.style_header_cell(ws.cell(r, 2 + i, value=header))

    first = r + 1
    for name, vals in rows:
        r += 1
        cc = ws.cell(r, 1, value=name)
        X.style_body_cell(cc)
        cc.alignment = X.LEFT
        for i, (key, _, fmt) in enumerate(DOD_COLUMNS):
            X.style_body_cell(ws.cell(r, 2 + i, value=vals.get(key, 0)), fmt)
    last = r

    # строка «Итог»
    r += 1
    cc = ws.cell(r, 1, value="Итог")
    X.style_body_cell(cc, bold=True)
    cc.fill = X.FILL_TOTAL
    cc.alignment = X.LEFT
    for i, (key, _, fmt) in enumerate(DOD_COLUMNS):
        cc = ws.cell(r, 2 + i, value=totals.get(key, 0))
        X.style_body_cell(cc, fmt, bold=True)
        cc.fill = X.FILL_TOTAL

    if last >= first:
        for i, (key, _, _fmt) in enumerate(DOD_COLUMNS):
            letter = X.col(2 + i)
            rng = f"{letter}{first}:{letter}{last}"
            # У ДРР «больше» значит «хуже»: обычная шкала красила самый
            # дорогой в рекламе товар зелёным, будто он лучший.
            if key == "drr":
                X.color_scale_inverted(ws, rng)
            else:
                X.color_scale(ws, rng)
    return r + 2


def _agg(rec, day_keys, has_views=True):
    """
    Свод метрик товара за набор дней + расчёт ДРР.

    has_views=False — OZON ещё не посчитал показы за эти дни (он отстаёт на
    сутки-двое). Тогда в колонке пусто, а не ноль: ноль рядом с непустыми
    продажами читается как «товар никто не видел».
    """
    v = {
        "revenue": round(P.sum_days(rec, day_keys, "revenue")),
        "ordered_units": int(round(P.sum_days(rec, day_keys, "ordered_units"))),
        "ad_spend": round(P.sum_days(rec, day_keys, "ad_spend")),
        "hits_view": (int(round(P.sum_days(rec, day_keys, "hits_view")))
                      if has_views else None),
    }
    v["drr"] = _safe_div(v["ad_spend"], v["revenue"])
    return v


def _totals(rows):
    t = {"revenue": 0, "ordered_units": 0, "ad_spend": 0, "hits_view": 0}
    known_views = False
    for _, v in rows:
        for k in t:
            val = v.get(k)
            if k == "hits_view":
                if val is None:
                    continue
                known_views = True
            t[k] += val or 0
    if not known_views:
        t["hits_view"] = None
    t["drr"] = _safe_div(t["ad_spend"], t["revenue"])
    return t


def _delta_rows(cur_rows, prev_rows):
    """Разности по товарам: ДРР считается как разница долей (в п.п.), как в образце."""
    cur = dict(cur_rows)
    prev = dict(prev_rows)
    out = []
    for name in list(cur) + [n for n in prev if n not in cur]:
        a, b = cur.get(name, {}), prev.get(name, {})
        d = {k: ((a.get(k) or 0) - (b.get(k) or 0))
             for k in ("revenue", "ordered_units", "ad_spend")}
        # разница показов имеет смысл, только если посчитаны оба дня
        av, bv = a.get("hits_view"), b.get("hits_view")
        d["hits_view"] = None if (av is None or bv is None) else av - bv
        d["drr"] = a.get("drr", 0) - b.get("drr", 0)
        out.append((name, d))
    return out


def _align_blocks(cur_rows, prev_rows):
    """
    Приводит оба блока к ОДНОМУ набору товаров в АЛФАВИТНОМ порядке — как в
    образце заказчика. Это принципиально: строки трёх блоков должны совпадать
    построчно, иначе их нельзя сравнивать глазами. Товар, которого нет в одном
    из периодов, добавляется с нулями.
    """
    empty = {"revenue": 0, "ordered_units": 0, "ad_spend": 0, "hits_view": 0, "drr": 0.0}
    cur, prev = dict(cur_rows), dict(prev_rows)
    names = sorted(set(cur) | set(prev), key=lambda s: str(s).lower())
    return ([(n, cur.get(n, dict(empty))) for n in names],
            [(n, prev.get(n, dict(empty))) for n in names])


def _build_dod_like(collectors, cfg, cur_from, cur_to, prev_from, prev_to,
                    lbl_cur, lbl_prev, lbl_delta, filename):
    wb = _new_wb()
    for colr in collectors:
        cur_daily, _ = colr.daily_by_product(cur_from, cur_to, only_in_stock=True)
        prev_daily, _ = colr.daily_by_product(prev_from, prev_to, only_in_stock=True)
        cur_keys = [D.d(x) for x in _daterange(cur_from, cur_to)]
        prev_keys = [D.d(x) for x in _daterange(prev_from, prev_to)]

        # за какие дни OZON вообще посчитал показы
        v_days = getattr(colr, "days_with_views", set()) or set()
        cur_v = any(k in v_days for k in cur_keys)
        prev_v = any(k in v_days for k in prev_keys)
        cur_rows = [(r["name"] or r["offer_id"], _agg(r, cur_keys, cur_v))
                    for r in cur_daily.values()]
        prev_rows = [(r["name"] or r["offer_id"], _agg(r, prev_keys, prev_v))
                     for r in prev_daily.values()]
        cur_rows, prev_rows = _align_blocks(cur_rows, prev_rows)

        ws = wb.create_sheet(X.safe_title(colr.name))
        r = 1
        r = _dod_block(ws, r, lbl_cur, cur_rows, _totals(cur_rows))
        r = _dod_block(ws, r, lbl_prev, prev_rows, _totals(prev_rows))
        deltas = _delta_rows(cur_rows, prev_rows)
        dt = _totals(deltas)
        dt["drr"] = _totals(cur_rows)["drr"] - _totals(prev_rows)["drr"]
        _dod_block(ws, r, lbl_delta, deltas, dt)

        X.set_widths(ws, [42, 12, 13, 12, 10, 12])
        ws.freeze_panes = "B2"
        X.page_setup(ws)

    path = _out(cfg, filename)
    wb.save(path)
    log.info("отчёт готов: %s", path)
    return path


def build_day_over_day(collectors, cfg):
    """Вчера к позавчера — макет образца «динамика день ко дню»."""
    tz = cfg.TIMEZONE
    y, dby = D.yesterday(tz), D.day_before_yesterday(tz)
    return _build_dod_like(
        collectors, cfg, y, y, dby, dby,
        lbl_cur=y.strftime("%d.%m.%Y"),
        lbl_prev=dby.strftime("%d.%m.%Y"),
        lbl_delta=f"динамика {y.strftime('%d.%m.')} к {dby.strftime('%d.%m.')}",
        filename=f"02_dinamika_den_ko_dnyu_{D.d(y)}.xlsx",
    )


def build_intraday(collectors, cfg, snapshots_dir=None):
    """
    Сегодня на время T к вчера на время T. Тот же макет, что отчёт 2.
    Использует снимки: каждый запуск сохраняет срез накопленного за день,
    сравнение идёт с вчерашним снимком того же часа.
    """
    snapshots_dir = snapshots_dir or getattr(cfg, "SNAPSHOTS_DIR", "snapshots")
    tz = cfg.TIMEZONE
    now = D.now_tz(tz)
    hour, today, yday = now.hour, now.date(), D.yesterday(tz)
    stamp = now.strftime("%Y-%m-%d_%H-%M")

    wb = _new_wb()
    for colr in collectors:
        cur_daily, _ = colr.daily_by_product(today, today, only_in_stock=True)
        cur_keys = [D.d(today)]
        v_days = getattr(colr, "days_with_views", set()) or set()
        # за сегодня показов у OZON заведомо ещё нет — колонка будет пустой
        cur_rows = [(r["name"] or r["offer_id"],
                     _agg(r, cur_keys, D.d(today) in v_days))
                    for r in cur_daily.values()]

        # снимок текущего слота — для сравнения завтра
        S.save(snapshots_dir, colr.name, D.d(today), hour,
               {n: v for n, v in cur_rows})

        snap = S.load(snapshots_dir, colr.name, D.d(yday), hour)
        if snap is not None:
            prev_rows = [(n, v) for n, v in snap.items()]
            note = ""
        else:
            prev_daily, _ = colr.daily_by_product(yday, yday, only_in_stock=True)
            prev_rows = [(r["name"] or r["offer_id"],
                          _agg(r, [D.d(yday)], D.d(yday) in v_days))
                         for r in prev_daily.values()]
            note = " (снимка за вчера на этот час нет — сравнение с полным вчерашним днём)"
        cur_rows, prev_rows = _align_blocks(cur_rows, prev_rows)

        ws = wb.create_sheet(X.safe_title(colr.name))
        r = 1
        r = _dod_block(ws, r, f"сегодня {today.strftime('%d.%m.')} на {hour:02d}:00",
                       cur_rows, _totals(cur_rows))
        r = _dod_block(ws, r, f"вчера {yday.strftime('%d.%m.')} на {hour:02d}:00{note}",
                       prev_rows, _totals(prev_rows))
        deltas = _delta_rows(cur_rows, prev_rows)
        dt = _totals(deltas)
        dt["drr"] = _totals(cur_rows)["drr"] - _totals(prev_rows)["drr"]
        _dod_block(ws, r, f"динамика на {hour:02d}:00", deltas, dt)

        X.set_widths(ws, [42, 12, 13, 12, 10, 12])
        ws.freeze_panes = "B2"
        X.page_setup(ws)

    path = _out(cfg, f"05_promezhutochnyy_{stamp}.xlsx")
    wb.save(path)
    log.info("отчёт 5 готов: %s", path)
    return path


# ============================================================ 3. Качественные

# Строки и форматы строго по образцу «по аналитике»:
# CTR — два знака (1,85%), % корзины и ДРР — один (20,5% / 14,7%),
# место в поиске — целое число (62, 72, 54).
QUALITY_ROWS = [
    # Подписи «(реклама)» стоят потому, что без выгрузки из кабинета клики
    # берутся только из рекламного отчёта: органических взять неоткуда, а
    # выдавать рекламные за общие нельзя. Появится выгрузка — подписи
    # снимаются, см. _quality_rows_for().
    ("показы", "hits_view", X.FMT_PLAIN_INT),
    ("клики (реклама)", "session_view", X.FMT_PLAIN_INT),
    ("CTR (реклама)", "ctr", X.FMT_PCT2),
    ("корзина", "hits_tocart", X.FMT_PLAIN_INT),
    ("% корзины", "cart_rate", X.FMT_PCT),
    ("купили (без отмен)", "bought", X.FMT_PLAIN_INT),
    ("отмен", "cancellations", X.FMT_PLAIN_INT),
    ("место в поиске", "position_category", X.FMT_PLAIN_INT),
    ("оборот", "revenue", X.FMT_PLAIN_INT),
    ("реклама", "ad_spend", X.FMT_PLAIN_INT),
    ("ДРР %", "drr", X.FMT_PCT),
]


def _quality_rows_for(filled):
    """
    Строки листа с учётом того, откуда пришли клики.

    Если выгрузка из кабинета закрыла session_view, клики в отчёте общие, а
    не рекламные, — пометка в названии становится враньём наоборот и её надо
    убрать. Пока выгрузки нет, пометка остаётся.
    """
    if "session_view" not in (filled or ()):
        return QUALITY_ROWS
    return [(label.replace(" (реклама)", ""), key, fmt)
            for label, key, fmt in QUALITY_ROWS]


def _quality_day_values(d, unified=False, has_views=True, has_cart=True):
    """
    Расчёт производных показателей за один день по образцу.

    unified=True — показы и клики пришли из одного источника (выгрузка
    кабинета), тогда CTR считается по ним. Иначе показы из поиска, а клики
    только рекламные, и делить одно на другое бессмысленно: CTR берётся по
    рекламной паре, а строка подписана «(реклама)».

    has_views / has_cart = False — за этот день источник данных не отдал
    НИЧЕГО. Тогда в клетке пусто, а не ноль. Разница принципиальная: OZON
    считает показы с задержкой в день-два, и на свежей колонке ноль читается
    как «товар никто не видел» — при том, что рядом в той же колонке стоят
    клики и оборот. Пустая клетка говорит правду: данных ещё нет.
    """
    views = d.get("hits_view", 0) or 0
    clicks = d.get("session_view", 0) or 0
    # CTR считаем от рекламной пары: показы приходят из поиска, клики — из
    # рекламы, и делить одно на другое было бы бессмыслицей.
    ad_views = d.get("ad_views", 0) or 0
    ad_clicks = d.get("ad_clicks", 0) or 0
    cart = d.get("hits_tocart", 0) or 0
    ordered = d.get("ordered_units", 0) or 0
    cancel = d.get("cancellations", 0) or 0
    revenue = d.get("revenue", 0) or 0
    spend = d.get("ad_spend", 0) or 0
    return {
        "hits_view": int(round(views)) if has_views else None,
        "session_view": int(round(clicks)),
        # рекламная пара нужна дальше, чтобы итог строки CTR считался тем же
        # способом, что и дни, и колонка «Итого» сходилась с колонками дней
        "ad_views": int(round(ad_views)),
        "ad_clicks": int(round(ad_clicks)),
        "ctr": ((_safe_div(clicks, views) if has_views else None) if unified
                else _safe_div(ad_clicks, ad_views)),
        "hits_tocart": int(round(cart)) if has_cart else None,
        "cart_rate": _safe_div(cart, clicks) if has_cart else None,
        "bought": int(round(max(ordered - cancel, 0))),
        "cancellations": int(round(cancel)),
        "position_category": (int(round(d.get("position_category", 0) or 0))
                              if has_views else None),
        "revenue": round(revenue),
        "ad_spend": round(spend),
        "drr": _safe_div(spend, revenue),
    }


# Строки, которые могут остаться пустыми не из-за магазина, а из-за того, что
# OZON не отдаёт данные. Клиент, увидев нули, решит, что товар никто не смотрел,
# поэтому под таблицей печатается пояснение — но только по тем строкам, которые
# в этом файле действительно вышли пустыми.
QUALITY_EMPTY_NOTES = [
    ("показы", "hits_view",
     "Их отдаёт метод «запросы моих товаров». Подписка Premium для него есть, "
     "но на этих аккаунтах он отвечает «нет данных за период» даже за дни, за "
     "которые данные у него заведомо есть: одно и то же окно то возвращает "
     "числа, то отказывает. Проверены формат дат, длина окна, недельное "
     "выравнивание и частота запросов — причина не найдена. Строка заполнится "
     "сама, как только метод начнёт отвечать: код запрашивает его каждый раз."),
    ("место в поиске", "position_category",
     "Тот же источник, что и «показы», — пусто по той же причине."),
    ("корзина", "hits_tocart",
     "По API её не отдаёт ни один метод без подписки Premium Plus. "
     "Заполняется выгрузкой из кабинета — см. ниже."),
    ("% корзины", "cart_rate",
     "Считается от «корзины», поэтому пусто вместе с ней."),
]

# Приписка под сноской: у всех четырёх строк одно и то же решение, и клиенту
# важнее знать, что делать, чем почему не работает.
QUALITY_EMPTY_FIX = (
    "Все эти строки заполнятся, если раз в день выгружать из кабинета отчёт "
    "«Аналитика → Графики» с разрезами «День» и «Товар» и класть файл в общую "
    "папку — инструкция в файле ВЫГРУЗКА_ИЗ_КАБИНЕТА.md. Второй вариант — "
    "подписка Premium Plus, тогда всё придёт по API само."
)


def _quality_empty_keys(store_totals, day_keys):
    """Какие из спорных строк вышли пустыми по всему магазину."""
    return [key for _label, key, _note in QUALITY_EMPTY_NOTES
            if not any(store_totals.get(k, {}).get(key) for k in day_keys)]


def _quality_write_notes(ws, r, empty_keys, width):
    """Сноска под таблицей. Пустых строк нет — ничего не пишем."""
    if not empty_keys:
        return r
    span = max(width, 2)
    r += 1
    X.style_header_cell(ws.cell(r, 1, value="Почему эти строки пустые"),
                        yellow=True)
    for i in range(1, span):
        X.style_header_cell(ws.cell(r, 1 + i, value=""))
    for label, key, note in QUALITY_EMPTY_NOTES:
        if key not in empty_keys:
            continue
        r += 1
        cc = ws.cell(r, 1, value=f"{label} — {note}")
        X.style_body_cell(cc)
        cc.alignment = X.LEFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    r += 1
    cc = ws.cell(r, 1, value=QUALITY_EMPTY_FIX)
    X.style_body_cell(cc, bold=True)
    cc.alignment = X.LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    return r + 1


def _quality_write_block(ws, r, title, days, vals_by_day, day_keys, with_total,
                         rows=None, unified=False):
    """Один блок «метрики строками, дни колонками». Возвращает следующую строку."""
    c = ws.cell(r, 1, value=title)
    X.style_header_cell(c, yellow=True)
    for i, dt in enumerate(days):
        c = ws.cell(r, 2 + i, value=dt)
        X.style_header_cell(c)
        c.number_format = X.FMT_DATE
    if with_total:
        X.style_header_cell(ws.cell(r, 2 + len(days), value="Итого"))

    for label, key, fmt in (rows or QUALITY_ROWS):
        r += 1
        cc = ws.cell(r, 1, value=label)
        X.style_body_cell(cc)
        cc.alignment = X.LEFT
        for i, k in enumerate(day_keys):
            X.style_body_cell(ws.cell(r, 2 + i, value=vals_by_day[k][key]), fmt)
        if with_total:
            X.style_body_cell(
                ws.cell(r, 2 + len(days),
                        value=_quality_row_total(key, vals_by_day, day_keys,
                                                 unified)),
                fmt, bold=True)
        # шкала по строке — как в образце, каждая метрика красится по своему ряду
        X.color_scale(ws, f"{X.col(2)}{r}:{X.col(1 + len(days))}{r}")
    return r + 2


def _quality_store_totals(items, day_keys, unified=False,
                          views_days=None, cart_days=None):
    """
    Свод по магазину на каждый день — блок «по аналитике».
    Аддитивные метрики складываются, «место в поиске» усредняется по товарам
    (складывать позиции в поиске бессмысленно — получилась бы сумма мест).

    Складываются УЖЕ ОКРУГЛЁННЫЕ значения товаров, а не сырые дроби. Иначе
    свод считается как round(сумма), а товарные блоки показывают
    сумму(round), и на полутора десятках товаров расходится на пару рублей.
    Клиент, сложивший колонку в Excel, увидит эти рубли и будет прав.
    """
    ADDITIVE = ("revenue", "ordered_units", "cancellations", "hits_view",
                "session_view", "hits_tocart", "ad_spend",
                "ad_views", "ad_clicks")
    out = {}
    for k in day_keys:
        acc = {m: 0 for m in ADDITIVE}
        positions = []
        for rec in items:
            day = rec["days"].get(k) or {}
            for m in ADDITIVE:
                acc[m] += round(day.get(m, 0) or 0)
            pos = day.get("position_category") or 0
            if pos:
                positions.append(pos)
        acc["position_category"] = (sum(positions) / len(positions)) if positions else 0
        out[k] = _quality_day_values(
            acc, unified,
            has_views=(views_days is None or k in views_days),
            has_cart=(cart_days is None or k in cart_days))
    return out


def build_quality(collectors, cfg):
    """
    Лист на магазин. Первый блок — свод «по аналитике» (как на образце),
    далее блок на каждый товар. Только товары на остатках, без OUT.
    """
    tz = cfg.TIMEZONE
    date_from, date_to = _period(cfg)
    days = _daterange(date_from, date_to)
    day_keys = [D.d(x) for x in days]
    stamp = D.now_tz(tz).strftime("%Y-%m-%d")
    with_total = bool(getattr(cfg, "QUALITY_TOTAL_COLUMN", False))

    wb = _new_wb()
    for colr in collectors:
        daily, _ = colr.daily_by_product(date_from, date_to, only_in_stock=True)
        items = sorted(
            daily.values(),
            key=lambda rec: P.sum_days(rec, day_keys, "revenue"),
            reverse=True,
        )

        ws = wb.create_sheet(X.safe_title(colr.name))
        r = 1

        # Клики из выгрузки кабинета — общие, а не рекламные: тогда и CTR
        # считается по своей паре, и пометка «(реклама)» из названий уходит.
        filled = getattr(colr, "cabinet_filled", set()) or set()
        unified = "session_view" in filled
        rows_def = _quality_rows_for(filled)
        # За какие дни источники вообще ответили. Дни вне этих множеств
        # выводятся пустыми, а не нулями.
        v_days = getattr(colr, "days_with_views", set()) or set()
        c_days = getattr(colr, "days_with_cart", set()) or set()

        def _vals(day_dict, k):
            return _quality_day_values(day_dict, unified,
                                       has_views=k in v_days,
                                       has_cart=k in c_days)

        # --- сводный блок по магазину ---
        empty_keys = []
        if items:
            store_totals = _quality_store_totals(items, day_keys, unified,
                                                 v_days, c_days)
            empty_keys = _quality_empty_keys(store_totals, day_keys)
            r = _quality_write_block(ws, r, "по аналитике", days,
                                     store_totals, day_keys, with_total,
                                     rows_def, unified)

        # --- блок на каждый товар ---
        for rec in items:
            vals_by_day = {k: _vals(rec["days"].get(k, {}), k) for k in day_keys}
            # Раньше условие смотрело только на показы и оборот, и товар,
            # на который лился бюджет без единой продажи, из отчёта исчезал —
            # при том что его расход входил в свод магазина. Это ровно тот
            # случай, который клиент и должен увидеть первым.
            if not any(v["hits_view"] or v["revenue"] or v["ad_spend"]
                       for v in vals_by_day.values()):
                continue
            r = _quality_write_block(ws, r, rec["name"] or rec["offer_id"], days,
                                     vals_by_day, day_keys, with_total,
                                     rows_def, unified)

        # --- сноска про пустые строки, если такие есть ---
        _quality_write_notes(ws, r, empty_keys,
                             1 + len(days) + (1 if with_total else 0))

        X.set_widths(ws, [24] + [11] * len(days) + ([12] if with_total else []))
        ws.freeze_panes = "B1"
        X.page_setup(ws)

    path = _out(cfg, f"03_kachestvennye_pokazateli_{stamp}.xlsx")
    wb.save(path)
    log.info("отчёт 3 готов: %s", path)
    return path


def _quality_row_total(key, vals_by_day, day_keys, unified=False):
    """Итог строки: аддитивные метрики суммируем, доли/позицию пересчитываем."""
    vals = [vals_by_day[k] for k in day_keys]
    if key in ("hits_view", "session_view", "hits_tocart", "bought",
               "cancellations", "revenue", "ad_spend"):
        known = [v[key] for v in vals if v[key] is not None]
        # все дни пустые — итог тоже пустой, а не бодрый ноль
        return round(sum(known)) if known else None
    if key == "ctr":
        # итог считается тем же способом, что и дни, иначе строка не сойдётся
        if unified:
            views = [v["hits_view"] for v in vals if v["hits_view"] is not None]
            if not views:
                return None
            return _safe_div(sum(v["session_view"] or 0 for v in vals),
                             sum(views))
        return _safe_div(sum(v.get("ad_clicks", 0) for v in vals),
                         sum(v.get("ad_views", 0) for v in vals))
    if key == "cart_rate":
        cart = [v["hits_tocart"] for v in vals if v["hits_tocart"] is not None]
        if not cart:
            return None
        return _safe_div(sum(cart), sum(v["session_view"] or 0 for v in vals))
    if key == "drr":
        return _safe_div(sum(v["ad_spend"] for v in vals), sum(v["revenue"] for v in vals))
    if key == "position_category":
        if all(v[key] is None for v in vals):
            return None
        nz = [v[key] for v in vals if v[key]]
        return round(sum(nz) / len(nz), 1) if nz else 0
    return 0


# ============================================================ 4. Остатки по кластерам

STOCK_HEADERS = [
    "Артикул", "Кластер", "Доступно к продаже", "В заявках на поставку",
    "В поставках в пути", "Итог", "прод 7д", "среднее", "ср/28 дней",
    "потреб 30д", "потреб 45д", "на сколько дней хватит остатков",
]


def build_stocks(collectors, cfg):
    """
    Артикул × кластер с прогнозом обеспеченности. Формулы — как в образце:
      Итог      = Доступно + В заявках + В пути
      среднее   = прод 7д / 7
      потреб 30д = среднее*30 - Итог
      потреб 45д = среднее*45 - Итог
      хватит дней = Итог / среднее

    ВСЁ СЧИТАЕТСЯ ПО КЛАСТЕРУ, а не по магазину. Раньше «прод 7д» и «среднее»
    брались из поля ads, которое OZON отдаёт «по всем кластерам», поэтому во
    всех строках товара стояло одно и то же число, а «потреб 30д» вычитала
    остаток одного кластера из потребности всего магазина. Заказчик это
    заметил, и он прав: смысла в такой цифре нет.

    Теперь источник — ads_cluster, среднесуточные продажи ИМЕННО в этом
    кластере за 28 дней. «прод 7д» — настоящие продажи товара за последнюю
    неделю, разложенные по кластерам пропорционально их доле в продажах
    (ads_cluster), а не по доле в остатках: остаток говорит о том, где товар
    лежит, а не где он продаётся. «ср/28 дней» — ads_cluster как есть.
    """
    tz = cfg.TIMEZONE
    stamp = D.now_tz(tz).strftime("%Y-%m-%d")
    date_to = D.yesterday(tz)
    date_from = date_to - timedelta(days=6)   # прод 7д

    wb = _new_wb()
    for colr in collectors:
        rows = colr.cluster_stocks()

        # Продажи за 7 дней по КЛАСТЕРУ ДОСТАВКИ — из выгрузки заказов, если
        # её положили. Это то, что просил заказчик: считать надо по тому, куда
        # товар уехал, а не откуда отгрузили. В API такого разреза нет вовсе.
        day_keys7 = set(D.d(x) for x in _daterange(date_from, date_to))
        orders = {}
        try:
            raw = getattr(colr, "cabinet_orders", lambda: {})() or {}
            for offer, by_cluster in raw.items():
                for cluster, days in by_cluster.items():
                    qty = sum(v for d, v in days.items() if d in day_keys7)
                    if qty:
                        orders.setdefault(offer, {})[cluster] = qty
            if orders:
                log.info("[%s] продажи по кластеру доставки из выгрузки заказов: "
                         "товаров %d", colr.name, len(orders))
        except Exception as e:
            log.warning("[%s] выгрузку заказов не разобрал: %s",
                        colr.name, str(e)[:200])

        # продажи за 7 дней по артикулу — запасной путь, если заказов нет
        sales7 = {}
        try:
            prods = colr.products_for_period(date_from, date_to, only_in_stock=True,
                                             with_kpi=False)
            for rec in prods.values():
                sales7[rec["offer_id"]] = int(round(rec.get("ordered_units", 0) or 0))
        except Exception as e:
            log.warning("[%s] продажи за 7 дней недоступны: %s", colr.name, e)

        ws = wb.create_sheet(X.safe_title(colr.name))
        # Шапка как в образце: A–E — серо-голубая шапка выгрузки OZON,
        # F «Итог» — без заливки, G–K (расчётные) — жёлтые с красным текстом.
        for i, h in enumerate(STOCK_HEADERS, start=1):
            cell = ws.cell(1, i, value=h)
            if i <= 5:
                X.style_header_cell(cell, ozon=True)
            elif i == 6:
                cell.border = X.BORDER
                cell.alignment = X.CENTER
            else:
                X.style_header_cell(cell, yellow=True)

        # Одна строка на кластер. OZON отдаёт строку на СКЛАД, а кластер
        # повторяется по нескольку раз — у «Москва, МО и Дальние регионы»
        # выходило девять строк на один товар. Заказчику нужен кластер целиком:
        # остатки складываем, а ads_cluster и idc_cluster НЕ складываем —
        # они и так посчитаны на весь кластер, во всех его строках одинаковы.
        merged = {}
        for rr in rows:
            key = (rr["offer_id"], rr.get("cluster", ""))
            cur = merged.get(key)
            if cur is None:
                merged[key] = dict(rr)
                continue
            for f in ("available", "requested", "transit"):
                cur[f] += rr.get(f, 0) or 0
            for f in ("ads", "idc", "ads_all"):
                if not cur.get(f):
                    cur[f] = rr.get(f) or 0
            if not cur.get("name"):
                cur["name"] = rr.get("name", "")

        # группируем по артикулу
        by_offer = {}
        for rr in merged.values():
            by_offer.setdefault(rr["offer_id"], []).append(rr)

        # считаем «прод 7д» заранее, чтобы отсортировать кластеры по нему
        prepared = []
        for offer_id, crows in by_offer.items():
            ads_sum = sum(rr.get("ads") or 0 for rr in crows)
            real7 = sales7.get(offer_id, 0)
            by_cluster = orders.get(offer_id) or {}
            total_stock = sum(rr["available"] + rr["requested"] + rr["transit"]
                              for rr in crows) or 1
            block = []
            for rr in crows:
                ads_c = rr.get("ads") or 0
                if by_cluster:
                    # лучший источник: сколько штук реально уехало в этот
                    # кластер за неделю, без всяких пропорций
                    sold7 = int(round(by_cluster.get(rr.get("cluster", ""), 0)))
                elif real7 and ads_sum:
                    # настоящая неделя, разложенная по доле кластера в продажах
                    sold7 = int(round(real7 * ads_c / ads_sum))
                elif ads_c:
                    # продаж за неделю нет под рукой — берём темп 28 дней
                    sold7 = int(round(ads_c * 7))
                else:
                    # OZON не дал продаж по кластеру: последнее средство —
                    # доля в остатках. Хуже, но лучше, чем ноль во всех строках
                    share = (rr["available"] + rr["requested"] + rr["transit"]) / total_stock
                    sold7 = int(round(real7 * share))
                block.append((rr, sold7, ads_c))
            # внутри артикула кластеры по убыванию продаж — как в образце
            block.sort(key=lambda t: t[1], reverse=True)
            prepared.append((offer_id, sum(x[1] for x in block), block))
        prepared.sort(key=lambda t: t[1], reverse=True)

        r = 1
        for offer_id, _, block in prepared:
            for rr, sold7, ads_c in block:
                r += 1
                # Именно артикул: наименования у заказчика по 120 символов,
                # и колонка становилась нечитаемой. Просил артикул.
                X.style_body_cell(ws.cell(r, 1, value=offer_id or rr.get("name", "")))
                ws.cell(r, 1).alignment = X.LEFT
                X.style_body_cell(ws.cell(r, 2, value=rr.get("cluster", "")))
                ws.cell(r, 2).alignment = X.LEFT
                X.style_body_cell(ws.cell(r, 3, value=rr["available"]), X.FMT_PLAIN_INT)
                X.style_body_cell(ws.cell(r, 4, value=rr["requested"]), X.FMT_PLAIN_INT)
                X.style_body_cell(ws.cell(r, 5, value=rr["transit"]), X.FMT_PLAIN_INT)

                # формулы 1-в-1 с образцом
                cc = ws.cell(r, 6, value=f"=C{r}+D{r}+E{r}")
                X.style_body_cell(cc); cc.number_format = X.FMT_PLAIN_INT
                X.style_body_cell(ws.cell(r, 7, value=sold7), X.FMT_PLAIN_INT)
                cc = ws.cell(r, 8, value=f"=G{r}/7")
                X.style_body_cell(cc); cc.number_format = X.FMT_FLOAT1
                # ср/28 дней — как есть от OZON, по этому кластеру
                X.style_body_cell(ws.cell(r, 9, value=round(ads_c, 1)),
                                  X.FMT_FLOAT1)
                cc = ws.cell(r, 10, value=f"=H{r}*30-F{r}")
                X.style_body_cell(cc); cc.number_format = X.FMT_PLAIN_INT
                cc = ws.cell(r, 11, value=f"=H{r}*45-F{r}")
                X.style_body_cell(cc); cc.number_format = X.FMT_PLAIN_INT
                cc = ws.cell(r, 12, value=f'=IF(H{r}=0,"",F{r}/H{r})')
                X.style_body_cell(cc); cc.number_format = X.FMT_FLOAT1

        if r > 1:
            # правила подсветки строго как в образце
            X.highlight_zero(ws, f"C2:E{r}")          # нет остатка — розовым
            X.highlight_zero(ws, f"F2:F{r}")          # нулевой итог — розовым
            X.highlight_negative_good(ws, f"J2:K{r}")  # потребность закрыта — зелёным
            X.color_scale(ws, f"G2:G{r}")             # прод 7д
            X.color_scale(ws, f"H2:H{r}")             # среднее
            X.color_scale(ws, f"I2:I{r}")             # ср/28 дней
            X.color_scale(ws, f"L2:L{r}")             # на сколько дней хватит

        X.set_widths(ws, [34, 30.6, 14.9, 14.9, 14.9, 9, 10, 10, 11, 11, 11, 14.9])
        ws.freeze_panes = "C2"
        X.page_setup(ws)

    path = _out(cfg, f"04_ostatki_po_pozitsiyam_{stamp}.xlsx")
    wb.save(path)
    log.info("отчёт 4 готов: %s", path)
    return path
