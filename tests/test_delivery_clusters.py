# -*- coding: utf-8 -*-
"""
Продажи за 7 дней по КЛАСТЕРУ ДОСТАВКИ из выгрузки заказов.

Заказчик считает потребность по тому, КУДА товар уехал, а не откуда его
отгрузили. В API кластера доставки нет ни в одном методе — единственный
источник — выгрузка заказов из кабинета. Здесь проверяется вся цепочка:
распознавание файла, разбор, слияние и подстановка в отчёт 4.
"""
import datetime as _dt
import io
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("OZON_STORES",
                      '[{"name":"ТЕСТ","client_id":"1","api_key":"k"}]')
os.environ.setdefault("DATA_DIR", "/tmp/_delivtest")

from openpyxl import Workbook, load_workbook

from ozon import cabinet as CAB
from ozon import dates as D
from ozon import reports as R

ok = True


def check(label, cond, extra=""):
    global ok
    print(("  OK   " if cond else "  ПРОВАЛ ") + label + ("" if cond else f"  <- {extra}"))
    ok = ok and cond


# Шапка ровно как в файле заказчика: «Кластер отгрузки» стоит ПЕРЕД
# «Кластером доставки», и перепутать их — та самая ошибка, на которую он
# пожаловался. «Наименование товара» стоит перед «Артикулом» нарочно:
# по подстроке «артикул» не должно поймать наименование.
HEAD = ["Номер заказа", "Принят в обработку", "Статус",
        "Наименование товара", "Артикул", "OZON id",
        "Количество", "Кластер отгрузки", "Кластер доставки"]


def order_row(day, status, name, offer, sku, qty, ship, deliver):
    return [1, day, status, name, offer, sku, qty, ship, deliver]


print("\n1. Файл заказов узнаётся по «Кластеру доставки»")
rows = [HEAD,
        order_row("05.08.2026", "Доставлен", "Мазь длинная", "ART-1", 111, 3,
                  "Хабаровск", "Москва")]
check("выгрузка заказов распознана", CAB.looks_like_orders(rows) is True)
check("аналитический файл заказами не считается",
      CAB.looks_like_orders([["Артикул", "Показы", "В корзину"],
                             ["ART-1", 10, 2]]) is False)
check("«кластер отгрузки» сам по себе — не признак",
      CAB.looks_like_orders([["Артикул", "Кластер отгрузки"], ["A", "Москва"]])
      is False)


print("\n2. Разбор: артикул -> кластер доставки -> день -> штуки")
rows = [HEAD,
        order_row("05.08.2026", "Доставлен", "Мазь", "ART-1", 111, 3,
                  "Хабаровск", "Москва"),
        order_row("05.08.2026", "Доставлен", "Мазь", "ART-1", 111, 2,
                  "Казань", "Москва"),
        order_row("06.08.2026", "Доставлен", "Мазь", "ART-1", 111, 4,
                  "Москва", "Ростов"),
        order_row("06.08.2026", "Отменён", "Мазь", "ART-1", 111, 99,
                  "Москва", "Москва"),
        order_row("06.08.2026", "Отменён покупателем", "Мазь", "ART-1", 111, 50,
                  "Москва", "Ростов"),
        order_row("07.08.2026", "Доставлен", "Крем", "ART-2", 222, 7,
                  "Москва", "Санкт-Петербург")]
out = CAB.parse_orders(rows, "тест")
check("товаров два", sorted(out) == ["ART-1", "ART-2"], sorted(out))
check("группировка по кластеру ДОСТАВКИ, а не отгрузки",
      sorted(out["ART-1"]) == ["Москва", "Ростов"], sorted(out["ART-1"]))
check("две строки одного дня и кластера сложились",
      out["ART-1"]["Москва"]["2026-08-05"] == 5, out["ART-1"]["Москва"])
check("отменённые не попали ни в один кластер",
      out["ART-1"]["Ростов"] == {"2026-08-06": 4}, out["ART-1"]["Ростов"])
check("день взят из «Принят в обработку»",
      list(out["ART-2"]["Санкт-Петербург"]) == ["2026-08-07"],
      out["ART-2"])
check("артикул, а не наименование", "Мазь" not in out, sorted(out))

print("\n3. Строки без кластера или без даты пропускаются молча")
rows_bad = [HEAD,
            order_row("", "Доставлен", "Мазь", "ART-1", 111, 3, "Москва", "Москва"),
            order_row("05.08.2026", "Доставлен", "Мазь", "ART-1", 111, 3, "Москва", ""),
            order_row("05.08.2026", "Доставлен", "Мазь", "ART-1", 111, 3,
                      "Москва", "Москва")]
out_bad = CAB.parse_orders(rows_bad, "тест")
check("осталась одна годная строка",
      out_bad == {"ART-1": {"Москва": {"2026-08-05": 3.0}}}, out_bad)

print("\n4. Файл без единой годной строки — понятная ошибка, а не мусор")
try:
    CAB.parse_orders([HEAD], "пустой")
    check("ошибка про пустые заказы", False, "исключения не было")
except CAB.CabinetImportError as e:
    check("ошибка про пустые заказы", "строк нет" in str(e), str(e))
try:
    CAB.parse_orders([["Кластер доставки"], ["Москва"]], "куцый")
    check("ошибка про недостающие колонки", False, "исключения не было")
except CAB.CabinetImportError as e:
    check("ошибка про недостающие колонки", "Кластер доставки" in str(e), str(e))


print("\n5. parse_file сам понимает, что ему дали")
def xlsx(rows, sheets=None):
    wb = Workbook()
    ws = wb.active
    ws.title = (sheets or ["Лист1"])[0]
    for r in rows:
        ws.append(r)
    for extra in (sheets or [])[1:]:
        wb.create_sheet(extra)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


kind, data = CAB.parse_file(xlsx([HEAD,
                                  order_row("05.08.2026", "Доставлен", "Мазь",
                                            "ART-1", 111, 3, "Казань", "Москва")]),
                            "заказы.xlsx")
check("тип orders", kind == "orders", kind)
check("данные на месте", data["ART-1"]["Москва"]["2026-08-05"] == 3, data)

kind2, data2 = CAB.parse_file(
    xlsx([["Артикул", "Уникальные посетители, всего", "В корзину"],
          ["ART-1", 100, 10]]), "2026-08-05.xlsx")
check("тип metrics", kind2 == "metrics", kind2)
check("метрики разобраны", data2["ART-1"]["2026-08-05"]["tocart"] == 10, data2)


print("\n6. Нужный лист ищется по шапке, а не берётся первым")
wb = Workbook()
wb.active.title = "Товар-склад"
wb["Товар-склад"].append(["Товар", "Склад", "Остаток"])
wb["Товар-склад"].append(["Мазь", "W1", 10])
ws2 = wb.create_sheet("продажи 7 дней убрала отмены ")
for r in [HEAD, order_row("05.08.2026", "Доставлен", "Мазь", "ART-1", 111, 3,
                          "Казань", "Москва")]:
    ws2.append(r)
buf = io.BytesIO()
wb.save(buf)
kind3, data3 = CAB.parse_file(buf.getvalue(), "рабочая книга.xlsx")
check("выбран лист с заказами, а не первый",
      kind3 == "orders" and data3["ART-1"]["Москва"]["2026-08-05"] == 3,
      (kind3, data3))

print("\n7. Миллион пустых строк в хвосте не съедает сбор")
wb = Workbook()
ws3 = wb.active
ws3.append(HEAD)
ws3.append(order_row("05.08.2026", "Доставлен", "Мазь", "ART-1", 111, 3,
                     "Казань", "Москва"))
ws3.cell(row=50000, column=9, value=None)   # растянутый лист, как у заказчика
buf = io.BytesIO()
wb.save(buf)
_t0 = _dt.datetime.now()
kind4, data4 = CAB.parse_file(buf.getvalue(), "хвост.xlsx")
_sec = (_dt.datetime.now() - _t0).total_seconds()
check("данные разобраны", data4["ART-1"]["Москва"]["2026-08-05"] == 3, data4)
check(f"чтение оборвалось на пустом хвосте ({_sec:.1f} с)", _sec < 10, _sec)


print("\n8. Слияние файлов: два уровня, артикул -> кластер")
dst = {"ART-1": {"Москва": {"2026-08-05": 5.0}}}
CAB._merge_orders(dst, {"ART-1": {"Москва": {"2026-08-06": 4.0},
                                  "Ростов": {"2026-08-05": 1.0}},
                        "ART-2": {"Казань": {"2026-08-05": 2.0}}})
check("новый день добавился к тому же кластеру",
      dst["ART-1"]["Москва"] == {"2026-08-05": 5.0, "2026-08-06": 4.0},
      dst["ART-1"]["Москва"])
check("новый кластер не затёр старый",
      sorted(dst["ART-1"]) == ["Москва", "Ростов"], sorted(dst["ART-1"]))
check("новый товар добавился", "ART-2" in dst, sorted(dst))


print("\n9. Папка магазина: заказы и аналитика лежат рядом")
FOLDER = "/tmp/_delivtest/import/ТЕСТ"
shutil.rmtree("/tmp/_delivtest", ignore_errors=True)
os.makedirs(FOLDER)
with open(os.path.join(FOLDER, "2026-08-05.xlsx"), "wb") as f:
    f.write(xlsx([["Артикул", "Уникальные посетители, всего", "В корзину"],
                  ["ART-1", 100, 10]]))
with open(os.path.join(FOLDER, "заказы.xlsx"), "wb") as f:
    f.write(xlsx([HEAD,
                  order_row("05.08.2026", "Доставлен", "Мазь", "ART-1", 111, 3,
                            "Казань", "Москва"),
                  order_row("06.08.2026", "Доставлен", "Мазь", "ART-1", 111, 4,
                            "Москва", "Ростов")]))
res = CAB.load_local("ТЕСТ", "/tmp/_delivtest")
check("метрики прочитаны", res["metrics"]["ART-1"]["2026-08-05"]["tocart"] == 10,
      res["metrics"])
check("заказы прочитаны из того же прогона",
      sorted(res["orders"]["ART-1"]) == ["Москва", "Ростов"], res["orders"])
check("пустая папка даёт обе половины пустыми",
      CAB.load_local("НЕТ ТАКОГО", "/tmp/_delivtest") == {"metrics": {}, "orders": {}},
      CAB.load_local("НЕТ ТАКОГО", "/tmp/_delivtest"))
shutil.rmtree("/tmp/_delivtest", ignore_errors=True)


print("\n10. Отчёт 4 берёт продажи из заказов, а не из пропорции")
# Остатки: почти всё лежит в Москве, продаётся тоже в Москве. Но по
# кластеру ДОСТАВКИ картина другая — и именно её надо показать.
CLUSTERS = [
    {"offer_id": "ART-1", "name": "Мазь", "cluster": "Москва", "warehouse": "W1",
     "available": 300, "requested": 0, "transit": 100, "ads": 30.0, "idc": 13.3},
    {"offer_id": "ART-1", "name": "Мазь", "cluster": "Ростов", "warehouse": "W2",
     "available": 50, "requested": 0, "transit": 0, "ads": 10.0, "idc": 5.0},
]
_to = D.yesterday("Europe/Moscow")
DAYS7 = [D.d(_to - _dt.timedelta(days=i)) for i in range(7)]
OLD = D.d(_to - _dt.timedelta(days=30))


class Store:
    name = "ТЕСТ"

    def __init__(self, orders=None, sales=280):
        self.orders = orders
        self.sales = sales

    def cluster_stocks(self):
        return CLUSTERS

    def cabinet_orders(self):
        if self.orders is None:
            raise RuntimeError("выгрузки нет")
        return self.orders

    def products_for_period(self, a, b, only_in_stock=True, with_kpi=True):
        return {"ART-1": {"offer_id": "ART-1", "ordered_units": self.sales}}


class Cfg:
    TIMEZONE = "Europe/Moscow"
    DATA_DIR = "/tmp/_delivtest"
    OUTPUT_DIR = "/tmp/_delivtest/out"


TOTAL = "все кластеры"


def build(store):
    """«прод 7д» по кластерам; итоговая строка позиции не в счёт."""
    ws = load_workbook(R.build_stocks([store], Cfg()))["ТЕСТ"]
    return {ws.cell(r, 2).value: ws.cell(r, 7).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 2).value and ws.cell(r, 2).value != TOTAL}


ORDERS = {"ART-1": {"Москва": {DAYS7[0]: 40.0, DAYS7[3]: 20.0},
                    "Ростов": {DAYS7[1]: 90.0}}}
res = build(Store(orders=ORDERS))
check("Москва — сумма своих дней из выгрузки", res["Москва"] == 60, res)
check("Ростов — свои 90, хотя ads_cluster у него втрое меньше",
      res["Ростов"] == 90, res)
check("пропорция по ads больше не применяется (было бы 210/70)",
      res != {"Москва": 210, "Ростов": 70}, res)

print("\n11. Считается ровно окно в 7 дней")
res = build(Store(orders={"ART-1": {"Москва": {DAYS7[6]: 5.0, OLD: 1000.0}}}))
check("седьмой день назад входит в окно", res["Москва"] == 5, res)
check("месячной давности заказ не входит", res["Москва"] != 1005, res)

print("\n12. Кластер, которого нет в выгрузке, получает ноль, а не пропорцию")
res = build(Store(orders={"ART-1": {"Москва": {DAYS7[0]: 60.0}}}))
check("Москва 60", res["Москва"] == 60, res)
check("Ростов 0 — туда за неделю ничего не уехало", res["Ростов"] == 0, res)

print("\n13. Нет выгрузки — работает прежний расчёт, отчёт не падает")
res = build(Store(orders=None))
check("вернулись к пропорции по ads_cluster",
      res == {"Москва": 210, "Ростов": 70}, res)
res = build(Store(orders={}))
check("пустая выгрузка тоже не ломает прежний путь",
      res == {"Москва": 210, "Ростов": 70}, res)


class Old:
    """Сборщик без cabinet_orders — на случай старого кода рядом."""
    name = "ТЕСТ"
    cluster_stocks = Store.cluster_stocks
    products_for_period = Store.products_for_period

    def __init__(self):
        self.sales = 280


check("сборщик без метода cabinet_orders не роняет отчёт",
      build(Old()) == {"Москва": 210, "Ростов": 70}, build(Old()))

shutil.rmtree("/tmp/_delivtest", ignore_errors=True)

print("\nИТОГ:", "все проверки пройдены" if ok else "ЕСТЬ ПРОВАЛЫ")
sys.exit(0 if ok else 1)
