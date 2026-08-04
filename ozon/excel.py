# -*- coding: utf-8 -*-
"""
Оформление Excel строго под образцы заказчика.

Из присланного «Отчет неделя_кол-во.xlsx» взято:
  * шапка блока: заливка + КРАСНЫЙ ЖИРНЫЙ шрифт;
  * ячейка с названием магазина — ЖЁЛТАЯ заливка (FFFF00), Arial 11, красный жирный;
  * даты в шапке — формат d-mmm;
  * трёхцветная шкала (красный F8696B → жёлтый FFEB84 → зелёный 63BE7B),
    min / 50-й перцентиль / max — накладывается на каждый блок отдельно;
  * тонкие границы по всем ячейкам таблицы.

Цвета вынесены в константы — при необходимости поменять оттенок правьте здесь.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

# --- палитра образца ---
YELLOW = "FFFF00"        # ячейка названия магазина / периода
HEADER_BG = "DDEBF7"     # шапка столбцов (даты, метрики)
RED_TEXT = "FFFF0000"    # цвет текста шапки
TOTAL_BG = "F2F2F2"      # строка «Итог»

SCALE_MIN = "F8696B"     # красный
SCALE_MID = "FFEB84"     # жёлтый
SCALE_MAX = "63BE7B"     # зелёный

# Подсветка отдельных значений (из образца «Распред мазь 5 ядов»):
ZERO_BG = "FFC7CE"       # ноль на остатках — розовая заливка
ZERO_TEXT = "9C0006"
GOOD_BG = "C6EFCE"       # отрицательная потребность (закупать не надо) — зелёная
GOOD_TEXT = "006100"

# Служебная шапка из выгрузки OZON (колонки «Артикул», «Кластер», «Доступно…»)
OZON_HEADER_BG = "F6F8FB"
OZON_HEADER_TEXT = "FF48525A"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_HEADER = Font(name="Arial", size=11, bold=True, color=RED_TEXT)
FONT_OZON_HEADER = Font(name="Calibri", size=11, bold=True, color=OZON_HEADER_TEXT)
FONT_NAME_CELL = Font(name="Calibri", size=11, bold=False)
FONT_TOTAL = Font(name="Calibri", size=11, bold=True)

FILL_YELLOW = PatternFill("solid", fgColor=YELLOW)
FILL_HEADER = PatternFill("solid", fgColor=HEADER_BG)
FILL_TOTAL = PatternFill("solid", fgColor=TOTAL_BG)
FILL_OZON_HEADER = PatternFill("solid", fgColor=OZON_HEADER_BG)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

FMT_INT = "#,##0"
FMT_PLAIN_INT = "0"       # как в выгрузке OZON — без разделителя разрядов
FMT_MONEY = "#,##0"
FMT_MONEY2 = "#,##0.00"
FMT_PCT = "0.0%"
FMT_PCT2 = "0.00%"        # CTR в образце с двумя знаками (1,85%)
FMT_FLOAT1 = "0.0"
FMT_FLOAT2 = "0.00"
FMT_DATE = "d-mmm"


def safe_title(title):
    """Excel запрещает : \\ / ? * [ ] и длину > 31."""
    for ch in ':\\/?*[]':
        title = str(title).replace(ch, " ")
    return title[:31] or "Лист"


def style_header_cell(cell, yellow=False, ozon=False):
    """
    yellow — жёлтая ячейка с красным жирным текстом (название магазина/периода
             и расчётные колонки в отчёте по остаткам);
    ozon   — серо-голубая шапка, как в выгрузке OZON («Артикул», «Кластер»…).
    """
    if ozon:
        cell.fill = FILL_OZON_HEADER
        cell.font = FONT_OZON_HEADER
    else:
        cell.fill = FILL_YELLOW if yellow else FILL_HEADER
        cell.font = FONT_HEADER
    cell.alignment = CENTER
    cell.border = BORDER


def style_body_cell(cell, number_format=None, bold=False):
    cell.border = BORDER
    cell.font = FONT_TOTAL if bold else FONT_NAME_CELL
    if number_format and isinstance(cell.value, (int, float)):
        cell.number_format = number_format


def color_scale(ws, cell_range):
    """Трёхцветная шкала как в образце (min / 50 перцентиль / max)."""
    ws.conditional_formatting.add(
        cell_range,
        ColorScaleRule(
            start_type="min", start_color=SCALE_MIN,
            mid_type="percentile", mid_value=50, mid_color=SCALE_MID,
            end_type="max", end_color=SCALE_MAX,
        ),
    )


def highlight_zero(ws, cell_range):
    """Ноль — розовая заливка с красным текстом (как в образце по остаткам)."""
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=["0"],
                   fill=PatternFill(start_color=ZERO_BG, end_color=ZERO_BG, fill_type="solid"),
                   font=Font(color=ZERO_TEXT)),
    )


def highlight_negative_good(ws, cell_range):
    """Отрицательное значение — зелёная заливка (потребность закрыта, закупать не надо)."""
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill(start_color=GOOD_BG, end_color=GOOD_BG, fill_type="solid"),
                   font=Font(color=GOOD_TEXT)),
    )


def set_widths(ws, widths):
    """widths: список чисел по столбцам, начиная с A."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def col(idx):
    return get_column_letter(idx)


def page_setup(ws, landscape=True):
    """Альбомная ориентация и подгонка по ширине — чтобы таблица не рвалась при печати."""
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = False
