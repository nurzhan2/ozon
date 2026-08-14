# -*- coding: utf-8 -*-
"""
Итоговая строка по позиции в отчёте 4 и внятные логи импорта из кабинета.

Заказчик (голосом): «добавить строчку, в которой будут указаны именно общие
все показатели по данной позиции — общий сток, общая заявка, в пути, общие
продажи за семь дней, средние общие. И далее вниз уже пошла разбивка по
кластерам».

И вторая половина: почему «корзина» осталась пустой. Источников у импорта
два, и оба умели молча ничего не делать — в логе не появлялось ни строки.
Здесь проверяется, что каждый прогон говорит, чем он воспользовался.
"""
import logging
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("OZON_STORES",
                      '[{"name":"ТЕСТ","client_id":"1","api_key":"k"}]')
os.environ.setdefault("DATA_DIR", "/tmp/_totals")

from openpyxl import load_workbook

from ozon import cabinet as CAB
from ozon import reports as R

ok = True


def check(label, cond, extra=""):
    global ok
    print(("  OK   " if cond else "  ПРОВАЛ ") + label + ("" if cond else f"  <- {extra}"))
    ok = ok and cond


TOTAL = "все кластеры"
CLUSTERS = [
    {"offer_id": "ART-1", "name": "Мазь", "cluster": "Москва", "warehouse": "W1",
     "available": 300, "requested": 0, "transit": 120, "ads": 30.0, "idc": 13.3},
    {"offer_id": "ART-1", "name": "Мазь", "cluster": "Ростов", "warehouse": "W2",
     "available": 73, "requested": 0, "transit": 0, "ads": 10.0, "idc": 5.0},
    {"offer_id": "ART-2", "name": "Тушь", "cluster": "Москва", "warehouse": "W3",
     "available": 50, "requested": 25, "transit": 0, "ads": 5.0, "idc": 2.0},
]


class Store:
    name = "ТЕСТ"

    def cluster_stocks(self):
        return CLUSTERS

    def cabinet_orders(self):
        return {}

    def products_for_period(self, a, b, only_in_stock=True, with_kpi=True):
        return {"ART-1": {"offer_id": "ART-1", "ordered_units": 280},
                "ART-2": {"offer_id": "ART-2", "ordered_units": 35}}


class Cfg:
    TIMEZONE = "Europe/Moscow"
    DATA_DIR = "/tmp/_totals"
    OUTPUT_DIR = "/tmp/_totals/out"
    GOOGLE_IMPORT_FOLDER = ""
    GOOGLE_CREDENTIALS_FILE = ""


shutil.rmtree("/tmp/_totals", ignore_errors=True)
ws = load_workbook(R.build_stocks([Store()], Cfg()))["ТЕСТ"]
rows = [(r, ws.cell(r, 1).value, ws.cell(r, 2).value)
        for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value]

print("\n1. Итоговая строка стоит НАД разбивкой по кластерам")
check("первая строка позиции — итоговая",
      rows[0][2] == TOTAL and rows[1][2] != TOTAL, rows[:3])
check("итоговых строк столько же, сколько позиций",
      sum(1 for _, _, c in rows if c == TOTAL) == 2,
      [c for _, _, c in rows])
check("в колонке «Артикул» итоговой строки — артикул",
      rows[0][1] == "ART-1", rows[0])
check("всего строк: две позиции + три кластера", len(rows) == 5, rows)

t1 = rows[0][0]
first, last = t1 + 1, t1 + 2

print("\n2. Общие остатки — сумма по кластерам, формулой")
check("сток", ws.cell(t1, 3).value == f"=SUM(C{first}:C{last})", ws.cell(t1, 3).value)
check("заявка", ws.cell(t1, 4).value == f"=SUM(D{first}:D{last})", ws.cell(t1, 4).value)
check("в пути", ws.cell(t1, 5).value == f"=SUM(E{first}:E{last})", ws.cell(t1, 5).value)
check("итог считается из своей же строки, как у кластеров",
      ws.cell(t1, 6).value == f"=C{t1}+D{t1}+E{t1}", ws.cell(t1, 6).value)

print("\n3. Общие продажи за 7 дней и средние")
check("прод 7д — сумма блока",
      ws.cell(t1, 7).value == f"=SUM(G{first}:G{last})", ws.cell(t1, 7).value)
check("среднее — своя седьмая часть",
      ws.cell(t1, 8).value == f"=G{t1}/7", ws.cell(t1, 8).value)
check("ср/28 дней — сумма кластерных",
      ws.cell(t1, 9).value == f"=SUM(I{first}:I{last})", ws.cell(t1, 9).value)
check("потребность 30д от своей строки",
      ws.cell(t1, 10).value == f"=H{t1}*30-F{t1}", ws.cell(t1, 10).value)
check("потребность 45д от своей строки",
      ws.cell(t1, 11).value == f"=H{t1}*45-F{t1}", ws.cell(t1, 11).value)
check("хватит дней от своей строки",
      ws.cell(t1, 12).value == f'=IF(H{t1}=0,"",F{t1}/H{t1})', ws.cell(t1, 12).value)

print("\n4. Диапазон SUM накрывает ровно свои кластеры")
# у второй позиции всего один кластер — диапазон из одной строки
t2 = [r for r, _, c in rows if c == TOTAL][1]
check("вторая позиция: SUM из одной строки",
      ws.cell(t2, 3).value == f"=SUM(C{t2 + 1}:C{t2 + 1})", ws.cell(t2, 3).value)
check("кластеры первой позиции в её диапазон не попали",
      f"C{t1}" not in str(ws.cell(t2, 3).value), ws.cell(t2, 3).value)

print("\n5. Итоговая строка заметна и не участвует в шкалах")
check("залита жёлтым, как в образце заказчика",
      ws.cell(t1, 1).fill.fgColor.rgb.endswith("FFFF00")
      or ws.cell(t1, 1).fill.fgColor.rgb != "00000000",
      ws.cell(t1, 1).fill.fgColor.rgb)
sqrefs = " ".join(str(rng.sqref) for rng in ws.conditional_formatting)
check("строка итога не попала ни в одну шкалу",
      f"G{t1}" not in sqrefs and f"C{t1}" not in sqrefs, sqrefs)
check("кластерные строки в шкалах остались",
      f"G{first}:G{last}" in sqrefs, sqrefs)
check("подсветка нуля не красит жёлтую строку",
      f"C{t2}" not in sqrefs, sqrefs)

print("\n6. Импорт из кабинета больше не молчит")


class Rec(logging.Handler):
    def __init__(self):
        super().__init__(logging.INFO)
        self.lines = []

    def emit(self, r):
        self.lines.append(r.getMessage())


def logs_of(cfg):
    h = Rec()
    lg = logging.getLogger("ozon.cabinet")
    lg.addHandler(h)
    lg.setLevel(logging.INFO)
    try:
        out = CAB.load("ТЕСТ", cfg)
    finally:
        lg.removeHandler(h)
    return out, "\n".join(h.lines)


class NoFolder(Cfg):
    pass


out, text = logs_of(NoFolder())
check("пустая переменная названа прямо",
      "GOOGLE_IMPORT_FOLDER" in text, text)
check("сказано, чем это грозит", "корзины" in text, text)
check("форма возврата правильная", out == {"metrics": {}, "orders": {}}, out)


class NoKey(Cfg):
    GOOGLE_IMPORT_FOLDER = "1abc"
    GOOGLE_CREDENTIALS_FILE = "/tmp/_totals/нет-такого.json"


out, text = logs_of(NoKey())
check("папка задана, а ключа нет — предупреждение",
      "ключа сервисного аккаунта нет" in text, text)
check("отчёт всё равно собирается",
      out == {"metrics": {}, "orders": {}}, out)


class LocalOnly(Cfg):
    DATA_DIR = "/tmp/_totals/пусто"


out, text = logs_of(LocalOnly())
check("про отсутствующую локальную папку тоже сказано",
      "локальной папки" in text, text)

print("\n7. Пустая подпапка на Диске возвращает правильную форму")
check("_empty(), а не голый {} — иначе вызывающий падал на KeyError",
      CAB._empty() == {"metrics": {}, "orders": {}}, CAB._empty())

shutil.rmtree("/tmp/_totals", ignore_errors=True)

print("\nИТОГ:", "все проверки пройдены" if ok else "ЕСТЬ ПРОВАЛЫ")
sys.exit(0 if ok else 1)
